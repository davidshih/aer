#!/usr/bin/env python3
"""
Stage 3: Reviewer Assignment with Complete Columns (v3.0 FINAL)

Complete output columns:
1. All input columns (preserved)
2. Reviewer (email from mapping)
3. Manual Review (Yes if email match, empty if dept match)
4. Reviewer's Response (dropdown: Approved/Denied/Changes Required)
5. Details of Access Change (free text for reviewer to fill)

Matching logic:
- Email exact match → Fill reviewer + Mark "Manual Review" = "Yes"
- Department contains match → Fill reviewer, no manual review mark
- Keep all input columns clean (no text pollution)

Usage in Jupyter:
    1. Run this cell
    2. Upload validated file from Stage 2
    3. Upload or use default mapping file
    4. Click "Assign Reviewers"
    5. Review assignments
    6. Click "Save Final Review"
"""

import os, sys, logging, glob, io
import pandas as pd
import ipywidgets as widgets
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from IPython.display import display, HTML, clear_output

# ============================================
# Setup Paths
# ============================================

today_str = datetime.now().strftime('%Y-%m-%d')
BASE_DIR = os.path.join("output", today_str)
STAGE2_DIR = os.path.join(BASE_DIR, "stage2_validated")
STAGE3_DIR = os.path.join(BASE_DIR, "stage3_review")
MAPPING_DIR = os.path.join("input", "mapping")
LOG_DIR = os.path.join(BASE_DIR, "logs")
os.makedirs(STAGE3_DIR, exist_ok=True)

# ============================================
# Logging
# ============================================

log_file = os.path.join(LOG_DIR, f"aer_stage3_{datetime.now().strftime('%Y%m%d_%H%M')}.log")
logger_s3 = logging.getLogger("aer_stage3")
logger_s3.handlers.clear()
logger_s3.setLevel(logging.INFO)
formatter = logging.Formatter('%(asctime)s | %(levelname)s | %(message)s')
fh = logging.FileHandler(log_file, encoding="utf-8")
fh.setFormatter(formatter)
logger_s3.addHandler(fh)
logger_s3.addHandler(logging.StreamHandler(sys.stdout))

# ============================================
# Global State
# ============================================

stage3_input_df = None
stage3_mapping_df = None
stage3_final_df = None
stage3_input_filename = ""

# ============================================
# UI Components
# ============================================

s3_upload = widgets.FileUpload(
    accept='.xlsx, .csv',
    description="Upload Validated File",
    button_style='info',
    layout=widgets.Layout(width='250px')
)
s3_upload_status = widgets.HTML(value="<i>No file selected</i>")

s3_upload_map = widgets.FileUpload(
    accept='.csv',
    description="Mapping File",
    button_style='info',
    layout=widgets.Layout(width='250px')
)
s3_map_status = widgets.HTML(value="<i>Will use default mapping if available</i>")

s3_btn_assign = widgets.Button(
    description="⚙️ Assign Reviewers",
    button_style='warning',
    layout=widgets.Layout(width='180px', height='40px'),
    disabled=True
)
s3_btn_save = widgets.Button(
    description="💾 Save Final Review",
    button_style='success',
    layout=widgets.Layout(width='180px', height='40px'),
    disabled=True
)
s3_status = widgets.HTML(value="<i>Please upload validated file from Stage 2</i>")
s3_output = widgets.Output()

# ============================================
# Helper Functions
# ============================================

def get_latest_mapping():
    """Get latest mapping file from default location"""
    try:
        files = glob.glob(os.path.join(MAPPING_DIR, "*.csv"))
        if not files:
            files = glob.glob(os.path.join(MAPPING_DIR, "**", "*.csv"), recursive=True)
        return max(files, key=os.path.getmtime) if files else None
    except:
        return None


def detect_column(df, candidates):
    """Detect column by name matching"""
    cols_lower = {str(c).lower().strip(): c for c in df.columns}

    for cand in candidates:
        for col_key, col_actual in cols_lower.items():
            if cand in col_key:
                return col_actual
    return None


def assign_reviewers(input_df, mapping_df):
    """
    Assign reviewers based on email/department mapping

    Args:
        input_df: Input DataFrame (must have 'Email', 'Department' columns)
        mapping_df: Mapping DataFrame with columns:
            - email: User email for exact matching (optional)
            - department: Department name/substring
            - reviewer: Reviewer email to assign

    Returns:
        DataFrame with:
        - All input columns (preserved)
        - Reviewer (email)
        - Manual Review (Yes if email match, empty otherwise)
        - Reviewer's Response (empty, for dropdown)
        - Details of Access Change (empty, for free text)
    """

    # Detect mapping columns
    map_cols = {
        'email': detect_column(mapping_df, ['email', 'mail']),
        'department': detect_column(mapping_df, ['department', 'dept']),
        'reviewer': detect_column(mapping_df, ['reviewer', 'owner', 'manager'])
    }

    if not map_cols['department'] or not map_cols['reviewer']:
        raise ValueError(
            f"Mapping file must have 'department' and 'reviewer' columns. "
            f"Found: {list(mapping_df.columns)}"
        )

    # Build mapping dictionaries
    email_map = {}
    dept_map = {}

    # Email mapping (if email column exists)
    if map_cols['email']:
        for _, row in mapping_df.iterrows():
            email_key = str(row[map_cols['email']]).lower().strip()
            if email_key and email_key not in ['nan', 'none', '*']:
                email_map[email_key] = str(row[map_cols['reviewer']]).strip()

    # Department mapping
    for _, row in mapping_df.iterrows():
        dept_key = str(row[map_cols['department']]).lower().strip()
        if dept_key and dept_key != 'nan':
            dept_map[dept_key] = str(row[map_cols['reviewer']]).strip()

    print(f"📊 Loaded mapping:")
    print(f"   Email mappings: {len(email_map)}")
    print(f"   Dept mappings: {len(dept_map)}\n")

    # Create output DataFrame (copy all input columns)
    output_df = input_df.copy()
    output_df['Reviewer'] = ''
    output_df['Manual Review'] = ''
    output_df['Reviewer\'s Response'] = ''
    output_df['Details of Access Change'] = ''

    # Statistics
    stats = {'email_match': 0, 'dept_match': 0, 'no_match': 0}

    # Process each user
    for idx, row in output_df.iterrows():
        user_email = str(row.get('Email', '')).lower().strip()
        user_dept = str(row.get('Department', '')).strip()

        reviewer_assigned = None
        is_manual_review = False

        # Priority 1: Email exact match → Manual Review = Yes
        if user_email and user_email != 'nan' and user_email in email_map:
            reviewer_assigned = email_map[user_email]
            is_manual_review = True
            stats['email_match'] += 1

        # Priority 2: Department contains match → Auto assign
        elif user_dept and user_dept not in ['N/A', 'nan', '']:
            user_dept_lower = user_dept.lower()

            # Try exact match first
            if user_dept_lower in dept_map:
                reviewer_assigned = dept_map[user_dept_lower]
                stats['dept_match'] += 1
            else:
                # Try contains logic
                # Example: input="corporate 123 - finance", mapping="finance" → MATCH
                for map_dept_key, map_reviewer in dept_map.items():
                    if map_dept_key in user_dept_lower:
                        reviewer_assigned = map_reviewer
                        stats['dept_match'] += 1
                        break

        # Assign results
        if reviewer_assigned:
            output_df.at[idx, 'Reviewer'] = reviewer_assigned
            output_df.at[idx, 'Manual Review'] = 'Yes' if is_manual_review else ''
        else:
            stats['no_match'] += 1

    # Print statistics
    print(f"📈 Assignment Results:")
    print(f"   Email Match (Manual Review): {stats['email_match']}")
    print(f"   Dept Match (Auto):           {stats['dept_match']}")
    print(f"   No Match:                    {stats['no_match']}")
    print(f"   Total:                       {len(output_df)}\n")

    return output_df


def _find_column_case_insensitive(df, candidates):
    """Find first matching column name by case-insensitive exact match."""
    lookup = {str(c).strip().lower(): c for c in df.columns}
    for cand in candidates:
        col = lookup.get(str(cand).strip().lower())
        if col:
            return col
    return None


def _is_yes(value) -> bool:
    return str(value).strip().lower() == 'yes'


def _is_inactive(value) -> bool:
    v = str(value).strip().lower()
    return v in {'no', 'false', '0', 'disabled', 'inactive'}


def apply_final_layout_and_sort(df):
    """
    Reorder and sort final output for easier manual review.

    Rules:
    1) First columns: Validation Status, is_AD_active, Manual Review
    2) Manual Review = Yes rows on top
    3) Auto-matched + inactive AD rows at bottom
    """
    out = df.copy()

    validation_col = _find_column_case_insensitive(out, ['Validation Status', 'Validation status'])
    active_col = _find_column_case_insensitive(out, ['is_AD_active', 'is_ad_active'])
    manual_col = _find_column_case_insensitive(out, ['Manual Review', 'manual_review', 'manu_review'])

    if not manual_col:
        out['Manual Review'] = ''
        manual_col = 'Manual Review'

    # Stable ordering anchor
    out['_orig_order'] = range(len(out))

    if validation_col is None:
        out['_validation_tmp'] = ''
        validation_col = '_validation_tmp'
    if active_col is None:
        out['_active_tmp'] = ''
        active_col = '_active_tmp'

    manual_yes = out[manual_col].map(_is_yes)
    inactive_ad = out[active_col].map(_is_inactive)
    auto_matched = out[validation_col].astype(str).str.lower().str.contains('auto', na=False)

    # 0 = top, 1 = middle, 2 = bottom
    out['_sort_bucket'] = 1
    out.loc[manual_yes, '_sort_bucket'] = 0
    out.loc[(~manual_yes) & inactive_ad & auto_matched, '_sort_bucket'] = 2

    out = out.sort_values(by=['_sort_bucket', '_orig_order'], ascending=[True, True], kind='stable')

    first_cols = []
    for col in [validation_col, active_col, manual_col]:
        if col in out.columns and col not in first_cols:
            first_cols.append(col)
    other_cols = [c for c in out.columns if c not in first_cols and not c.startswith('_')]
    out = out[first_cols + other_cols]

    drop_cols = [c for c in out.columns if c.startswith('_')]
    if drop_cols:
        out = out.drop(columns=drop_cols, errors='ignore')

    return out


def save_with_validation(df, output_path):
    """
    Save DataFrame to Excel with data validation for dropdown columns

    Dropdowns:
    - Manual Review: Yes/No
    - Reviewer's Response: Approved/Denied/Changes Required
    """

    # Save to Excel
    df.to_excel(output_path, index=False, sheet_name='Review')

    # Add data validation
    wb = load_workbook(output_path)
    ws = wb.active

    # Find columns
    col_indices = {}
    for i, col_name in enumerate(df.columns, 1):
        col_str = str(col_name)
        if col_str == 'Manual Review':
            col_indices['manual_review'] = i
        elif col_str == 'Reviewer\'s Response':
            col_indices['reviewer_response'] = i

    # Manual Review dropdown (Yes/No)
    if 'manual_review' in col_indices:
        dv = DataValidation(
            type="list",
            formula1='"Yes,No"',
            allow_blank=True
        )
        dv.promptTitle = "Manual Review Status"
        dv.prompt = "Select Yes if this assignment requires manual review"
        dv.showInputMessage = True

        ws.add_data_validation(dv)
        letter = get_column_letter(col_indices['manual_review'])
        dv.add(f"{letter}2:{letter}{len(df)+1}")

    # Reviewer's Response dropdown (Approved/Denied/Changes Required)
    if 'reviewer_response' in col_indices:
        dv = DataValidation(
            type="list",
            formula1='"Approved,Denied,Changes Required"',
            allow_blank=True
        )
        dv.promptTitle = "Reviewer's Response"
        dv.prompt = "Select the review decision"
        dv.showInputMessage = True
        dv.errorTitle = "Invalid Selection"
        dv.error = "Please select a valid option from the dropdown"
        dv.showErrorMessage = True

        ws.add_data_validation(dv)
        letter = get_column_letter(col_indices['reviewer_response'])
        dv.add(f"{letter}2:{letter}{len(df)+1}")

    # Auto-fit column widths and enable wrap text for readability
    for col_idx, col_name in enumerate(df.columns, 1):
        letter = get_column_letter(col_idx)
        col_values = [str(col_name)]
        for val in df.iloc[:, col_idx - 1]:
            if pd.isna(val):
                continue
            col_values.append(str(val))

        max_len = max((len(v) for v in col_values), default=10)
        width = min(max(12, max_len + 2), 60)
        ws.column_dimensions[letter].width = width

        for row_idx in range(2, len(df) + 2):
            cell = ws[f"{letter}{row_idx}"]
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    wb.save(output_path)
    print(f"✅ Added dropdown validation for:")
    if 'manual_review' in col_indices:
        print(f"   - Manual Review (Yes/No)")
    if 'reviewer_response' in col_indices:
        print(f"   - Reviewer's Response (Approved/Denied/Changes Required)")
    print()

# ============================================
# Event Handlers
# ============================================

def on_s3_upload_change(change):
    """Handle input file upload"""
    if s3_upload.value and len(s3_upload.value) > 0:
        fname = s3_upload.value[0]['name']
        s3_upload_status.value = f"<b style='color:green;'>✅ Selected: {fname}</b>"
        s3_btn_assign.disabled = False


def on_s3_map_change(change):
    """Handle mapping file upload"""
    if s3_upload_map.value and len(s3_upload_map.value) > 0:
        fname = s3_upload_map.value[0]['name']
        s3_map_status.value = f"<b style='color:green;'>✅ Using: {fname}</b>"


def do_stage3_assign(b):
    """Assign reviewers based on mapping"""
    global stage3_input_df, stage3_mapping_df, stage3_final_df, stage3_input_filename

    s3_output.clear_output()

    if not s3_upload.value:
        with s3_output:
            print("❌ Please upload a validated file from Stage 2")
        return

    b.disabled = True

    try:
        with s3_output:
            print("\n" + "="*60)
            print("⚙️ Stage 3: Reviewer Assignment (Complete)")
            print("="*60 + "\n")

        # Load input file
        f_item = s3_upload.value[0]
        stage3_input_filename = f_item['name']

        if stage3_input_filename.endswith('.csv'):
            stage3_input_df = pd.read_csv(io.BytesIO(f_item['content']))
        else:
            stage3_input_df = pd.read_excel(io.BytesIO(f_item['content']))

        logger_s3.info(f"Loaded input: {stage3_input_filename}, {len(stage3_input_df)} rows")

        with s3_output:
            print(f"📄 Input file: {stage3_input_filename}")
            print(f"   Rows: {len(stage3_input_df)}")
            print(f"   Columns: {list(stage3_input_df.columns)}\n")

        # Validate required columns
        if 'Email' not in stage3_input_df.columns:
            with s3_output:
                print(f"❌ Input file missing 'Email' column")
                print(f"   Available: {list(stage3_input_df.columns)}")
            b.disabled = False
            return

        if 'Department' not in stage3_input_df.columns:
            with s3_output:
                print("⚠️  Warning: 'Department' column not found. Adding placeholder.")
            stage3_input_df['Department'] = 'N/A'

        # Load mapping file
        map_src = None
        map_name = ""

        if s3_upload_map.value and len(s3_upload_map.value) > 0:
            map_src = io.BytesIO(s3_upload_map.value[0]['content'])
            map_name = s3_upload_map.value[0]['name']
        else:
            default_map = get_latest_mapping()
            if default_map:
                map_src = default_map
                map_name = os.path.basename(default_map)

        if not map_src:
            with s3_output:
                print("❌ No mapping file found")
                print(f"   Upload a mapping file or place one in: {MAPPING_DIR}")
            s3_status.value = "<span style='color:red;'>❌ Mapping file required</span>"
            b.disabled = False
            return

        stage3_mapping_df = pd.read_csv(map_src)

        with s3_output:
            print(f"📋 Mapping file: {map_name}")
            print(f"   Rows: {len(stage3_mapping_df)}")
            print(f"   Columns: {list(stage3_mapping_df.columns)}\n")

        logger_s3.info(f"Loaded mapping: {map_name}, {len(stage3_mapping_df)} rows")

        # Assign reviewers
        with s3_output:
            print("="*60)
            print("🔍 Assigning Reviewers")
            print("="*60 + "\n")

        stage3_final_df = assign_reviewers(stage3_input_df, stage3_mapping_df)

        # Preview results
        with s3_output:
            print("="*60)
            print("📊 Preview (first 5 rows)")
            print("="*60)
            preview_cols = ['Email', 'Department', 'Reviewer', 'Manual Review']
            available_cols = [c for c in preview_cols if c in stage3_final_df.columns]
            print(stage3_final_df[available_cols].head())
            print()

        s3_btn_save.disabled = False
        s3_status.value = f"<span style='color:green;'>✅ Assignment complete: {len(stage3_final_df)} users ready</span>"
        logger_s3.info(f"Assignment complete: {len(stage3_final_df)} rows")

    except Exception as e:
        with s3_output:
            print(f"\n❌ Error: {str(e)}")
        logger_s3.error(f"Assignment error: {str(e)}", exc_info=True)
        s3_status.value = f"<span style='color:red;'>❌ Error: {str(e)}</span>"
    finally:
        b.disabled = False


def do_stage3_save(b):
    """Save final review file with dropdowns"""
    if stage3_final_df is None:
        with s3_output:
            print("❌ No data to save. Please assign reviewers first.")
        return

    b.disabled = True

    try:
        with s3_output:
            print("\n" + "="*60)
            print("💾 Saving Final Review File")
            print("="*60 + "\n")

        # Generate filename
        base_name = stage3_input_filename.replace('.csv', '').replace('.xlsx', '').replace('_AD_verified', '')
        timestamp = datetime.now().strftime('%Y%m%d_%H%M')
        output_filename = f"{base_name}_review_{timestamp}.xlsx"
        output_path = os.path.join(STAGE3_DIR, output_filename)

        # Apply final sort/layout before save
        final_df = apply_final_layout_and_sort(stage3_final_df)

        # Save with validation + formatting
        save_with_validation(final_df, output_path)

        with s3_output:
            print(f"✅ Saved to: {output_path}")
            print(f"   Rows: {len(final_df)}")
            print(f"   Columns: {list(final_df.columns)}")
            print("="*60)

        s3_status.value = f"<span style='color:blue;'>✅ Saved: {output_filename}</span>"
        logger_s3.info(f"Saved review file: {output_path}")

    except Exception as e:
        with s3_output:
            print(f"\n❌ Save error: {str(e)}")
        logger_s3.error(f"Save error: {str(e)}", exc_info=True)
        s3_status.value = f"<span style='color:red;'>❌ Save error: {str(e)}</span>"
    finally:
        b.disabled = False

# ============================================
# Bind Events
# ============================================

s3_upload.observe(on_s3_upload_change, 'value')
s3_upload_map.observe(on_s3_map_change, 'value')
s3_btn_assign.on_click(do_stage3_assign)
s3_btn_save.on_click(do_stage3_save)

# ============================================
# Initialize: Check for default mapping
# ============================================

default_map = get_latest_mapping()
if default_map:
    s3_map_status.value = f"<b style='color:green;'>✅ Default: {os.path.basename(default_map)}</b>"
    logger_s3.info(f"Default mapping found: {os.path.basename(default_map)}")
else:
    s3_map_status.value = f"<i>No default mapping found in {MAPPING_DIR}</i>"
    logger_s3.warning(f"No default mapping found")

# ============================================
# UI Layout
# ============================================

stage3_ui = widgets.VBox([
    widgets.HTML("""
        <div style='
            background: linear-gradient(135deg, #fa709a 0%, #fee140 100%);
            padding: 20px;
            border-radius: 8px;
            color: white;
            margin-bottom: 20px;
            box-shadow: 0 4px 6px rgba(0,0,0,0.1);
        '>
            <h2 style='margin: 0 0 10px 0;'>⚙️ Stage 3: Reviewer Assignment (Complete)</h2>
            <p style='margin: 0; opacity: 0.9;'>
                Assign reviewers + prepare complete review sheet with all required columns
            </p>
            <div style='margin-top: 10px; padding: 10px; background: rgba(255,255,255,0.2); border-radius: 5px;'>
                <b>Output columns:</b> Reviewer | Manual Review | Reviewer's Response | Details of Access Change
            </div>
        </div>
    """),
    widgets.HBox([s3_upload, s3_upload_status]),
    widgets.HBox([s3_upload_map, s3_map_status]),
    widgets.HBox([s3_btn_assign, s3_btn_save], layout=widgets.Layout(margin='10px 0')),
    s3_status,
    s3_output
])

clear_output()
display(stage3_ui)

logger_s3.info("Stage 3 UI initialized (Complete version)")
logger_s3.info("="*60)
