# === Cell 5: AER Engine v4.2 (Cache Stats Fix & Global Report) ===
import pandas as pd
import re, time, json, os, requests
import ipywidgets as widgets
from datetime import datetime
from openpyxl import load_workbook
from io import BytesIO
from IPython.display import display, HTML, clear_output

# ==========================================
# PART 1: CONFIG & LOADER
# ==========================================
CACHE_FILE = "aer_cache.json"
NOTES_FILE = "aer_manual_notes.json"
CACHE_VERSION = 4.2
cache_updated = False
EXCLUDED_FOLDERS = ["forms", "_private", "user listings", "audit logs", "audit"]

def load_json(file_path):
    if os.path.exists(file_path):
        try:
            with open(file_path, 'r', encoding='utf-8') as f: return json.load(f)
        except: return {}
    return {}

def save_json(file_path, data):
    with open(file_path, 'w', encoding='utf-8') as f: json.dump(data, f, indent=2, ensure_ascii=False)

local_cache = load_json(CACHE_FILE)
manual_data_store = load_json(NOTES_FILE)

# ==========================================
# PART 2: ROBUST EXCEL PARSER
# ==========================================
def find_col_index(headers, keywords):
    """
    Case-insensitive search for column index.
    """
    for idx, h in enumerate(headers):
        if not h: continue
        h_str = str(h).strip().lower()
        if all(k in h_str for k in keywords):
            return idx
    return None

def read_excel_rows(excel_bytes: bytes, reviewer_name: str, file_name: str, folder_url: str) -> list[dict]:
    wb = load_workbook(BytesIO(excel_bytes), read_only=True)

    # 1. Smart Tab Selection
    sheet_name = wb.sheetnames[0]
    for sn in wb.sheetnames:
        if "user listing" in sn.lower(): sheet_name = sn; break
    ws = wb[sheet_name]

    # 2. Read Headers
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        return []

    # 3. Robust Column Mapping (Fuzzy Match)
    idx_rev = find_col_index(header, ["reviewer"])
    idx_res = find_col_index(header, ["response"])
    idx_det = find_col_index(header, ["details", "change"])

    # User being reviewed: name & email
    idx_user = find_col_index(header, ["user", "name"])
    if idx_user is None: idx_user = find_col_index(header, ["display", "name"])
    if idx_user is None: idx_user = find_col_index(header, ["full", "name"])
    idx_email = find_col_index(header, ["email"])
    if idx_email is None: idx_email = find_col_index(header, ["mail"])

    # Validation
    if idx_rev is None or idx_res is None:
        if idx_rev is None: idx_rev = find_col_index(header, ["manager"]) # Fallback
        if idx_rev is None or idx_res is None: return []

    results = []

    # 4. Iterate Rows
    for i, row in enumerate(rows_iter, start=2):
        r_rev = row[idx_rev] if idx_rev < len(row) else None
        r_res = row[idx_res] if idx_res < len(row) else None
        r_det = row[idx_det] if idx_det is not None and idx_det < len(row) else None
        r_user = row[idx_user] if idx_user is not None and idx_user < len(row) else None
        r_email = row[idx_email] if idx_email is not None and idx_email < len(row) else None

        # Filter by Reviewer Name
        if str(r_rev).strip().lower() != reviewer_name.lower(): continue

        # Clean Response
        val_res = str(r_res).strip() if r_res else ""
        val_det = str(r_det).strip() if r_det else ""

        results.append({
            "reviewer": reviewer_name,
            "user_name": str(r_user).strip() if r_user else "",
            "user_email": str(r_email).strip() if r_email else "",
            "response": val_res,
            "details": val_det,
            "is_missing": (val_res == ""),
            "row_number": i,
            "file_name": file_name,
            "folder_url": folder_url
        })
    return results

def get_row_stats(txt):
    txt = str(txt).lower().strip()
    kw_appr = ['approv', 'retain', 'keep', 'confirm', 'yes', 'ok', 'active']
    kw_deny = ['denied', 'deny', 'remove', 'delete', 'revok', 'reject', 'no']
    kw_chg  = ['change', 'modif', 'updat', 'correct', 'edit', 'adjust']
    return {
        "is_appr": 1 if any(k in txt for k in kw_appr) else 0,
        "is_deny": 1 if any(k in txt for k in kw_deny) else 0,
        "is_chg":  1 if any(k in txt for k in kw_chg) else 0
    }

# ==========================================
# PART 3: SCANNING ENGINE
# ==========================================
if 'TARGET_APPS' not in globals() or not TARGET_APPS:
    print("⚠️ Please select Apps in Cell 2 first!")
    TARGET_APPS = []

all_responses = []
errors = []

print(f"🚀 Starting Scan Engine v{CACHE_VERSION} (Fuzzy Column Match)...")

for category, current_app_name, current_path in TARGET_APPS:
    try:
        raw_folders = list_folders_with_count(site_id, current_path) if 'list_folders_with_count' in globals() else []
        if not raw_folders: raw_folders = [{"name": "Error", "webUrl": "#"}]

        reviewers = [r for r in raw_folders if r['name'].lower() not in EXCLUDED_FOLDERS]
        total_revs = len(reviewers)
        logger.info(f"📂 App: {current_app_name} | Reviewers: {total_revs}")

        for idx, folder in enumerate(reviewers, 1):
            reviewer_name = folder["name"]
            folder_url = folder["webUrl"]
            folder_path = f"{current_path}/{reviewer_name}"
            cache_key = f"{category}|{current_app_name}|{reviewer_name}"

            try:
                # 1. File Discovery
                files = list_excel_files(site_id, folder_path)
                match_candidates = [f for f in files if reviewer_name.lower() in f["name"].lower()]
                target_file = match_candidates[0] if match_candidates else (files[0] if files else None)

                if not target_file:
                    continue

                remote_mod = target_file.get("lastModifiedDateTime")

                # 2. Cache Check (Version Controlled)
                cached = local_cache.get(cache_key)
                is_hit = False

                if USE_CACHE and cached and cached.get('v') == CACHE_VERSION and cached.get('last_mod') == remote_mod:
                    if 'rows' in cached and len(cached['rows']) > 0: is_hit = True

                if is_hit:
                    audit_snap = cached.get('audit', {})
                    c_appr = cached['stats']['Appr']
                    c_deny = cached['stats']['Deny']
                    c_chg = cached['stats']['Chg']
                    c_miss = sum(1 for r in cached['rows'] if r['is_missing'])

                    logger.info(f"  ✅ Read (Cache): [{idx}/{total_revs}] {reviewer_name} (Missing:{c_miss})(A:{c_appr}, D:{c_deny}, C:{c_chg})")

                    for r in cached['rows']:
                        r_copy = r.copy()
                        st = get_row_stats(r['response'])
                        r_copy.update({
                            "Category": category, "App_Name": current_app_name,
                            "Last_Modified": remote_mod, "File_Created_Date": audit_snap.get('created_ts'),
                            "Audit_Log": audit_snap.get('log'), "File_Creator": audit_snap.get('creator'), "File_Modifier": audit_snap.get('modifier'),
                            "stats_appr": st['is_appr'], "stats_deny": st['is_deny'], "stats_chg": st['is_chg'],
                            "source_is_cache": True
                        })
                        all_responses.append(r_copy)
                    continue

                # 3. Process Live File
                content = download_file(site_id, f"{folder_path}/{target_file['name']}")
                audit_info = get_file_audit_info(site_id, target_file["id"])

                rows = read_excel_rows(content, reviewer_name, target_file['name'], folder_url)

                s_appr, s_deny, s_chg, miss_cnt = 0, 0, 0, 0
                clean_rows_cache = []
                final_created = audit_info.get('created_ts') or target_file.get("createdDateTime")

                for r in rows:
                    st = get_row_stats(r['response'])
                    s_appr += st['is_appr']; s_deny += st['is_deny']; s_chg += st['is_chg']
                    if r['is_missing']: miss_cnt += 1

                    clean_rows_cache.append({
                        "reviewer": r['reviewer'], "user_name": r.get('user_name', ''), "user_email": r.get('user_email', ''),
                        "response": r['response'], "details": r['details'],
                        "is_missing": r['is_missing'], "row_number": r['row_number'],
                        "file_name": r['file_name'], "folder_url": r['folder_url']
                    })

                    r.update({
                        "Category": category, "App_Name": current_app_name,
                        "Last_Modified": remote_mod, "File_Created_Date": final_created,
                        "Audit_Log": audit_info['log'], "File_Creator": audit_info['creator'], "File_Modifier": audit_info['modifier'],
                        "stats_appr": st['is_appr'], "stats_deny": st['is_deny'], "stats_chg": st['is_chg'],
                        "source_is_cache": False
                    })

                all_responses.extend(rows)

                logger.info(f"  ✅ Read: [{idx}/{total_revs}] {reviewer_name} (Missing:{miss_cnt})(A:{s_appr}, D:{s_deny}, C:{s_chg})")

                if rows:
                    local_cache[cache_key] = {
                        "v": CACHE_VERSION, "last_mod": remote_mod,
                        "stats": {"Appr": s_appr, "Deny": s_deny, "Chg": s_chg},
                        "audit": audit_info, "rows": clean_rows_cache
                    }
                    cache_updated = True

            except Exception as e:
                logger.error(f"  ❌ Error {reviewer_name}: {e}")
                errors.append({"Category": category, "App_Name": current_app_name, "reviewer": reviewer_name, "error": str(e), "folder_url": folder_url})

    except Exception as e: logger.error(f"❌ App Error: {e}")

if cache_updated:
    save_json(CACHE_FILE, local_cache)
    logger.info("💾 Cache Updated (v4.2)")

# ==========================================
# PART 4: DASHBOARD
# ==========================================
df = pd.DataFrame(all_responses)
widget_store = {}
unified_data = {}
today_str = datetime.now().strftime("%Y-%m-%d")
output_dir = f"output/{today_str}"
os.makedirs(output_dir, exist_ok=True)

if not df.empty:
    stats = df.groupby(["Category", "App_Name", "reviewer"]).agg(
        missing=("is_missing", "sum"),
        approved=("stats_appr", "sum"), denied=("stats_deny", "sum"), changed=("stats_chg", "sum"),
        f_creator=("File_Creator", "first"), f_modifier=("File_Modifier", "first"), audit=("Audit_Log", "first"),
        is_cached=("source_is_cache", "all")
    ).reset_index()

    for _, row in stats.iterrows():
        key = f"{row['Category']} > {row['App_Name']}"
        if key not in unified_data:
            saved_app = manual_data_store.get(key, {})
            unified_data[key] = {
                "Category": row['Category'], "App_Name": row['App_Name'],
                "status_manual": saved_app.get("app_status", "Calculated"), "note_manual": saved_app.get("app_note", ""),
                "reviewers": {}, "stats": {"total_users": 0, "completed_users": 0}
            }

        node = unified_data[key]
        node['stats']['total_users'] += 1
        is_done = (row['missing'] == 0)
        if is_done: node['stats']['completed_users'] += 1
        status_calc = f"❌ Pending: {row['missing']}"
        if is_done: status_calc = "✅ Cached - Completed" if row['is_cached'] else "✅ Completed"

        d_style = "color:red;font-weight:bold" if row['denied'] > 0 else "color:#555"
        node['reviewers'][row['reviewer']] = {
            "status_calc": status_calc,
            "detail_html": f"Appr:{int(row['approved'])} | <span style='{d_style}'>Deny:{int(row['denied'])}</span> | Chg:{int(row['changed'])}",
            "folder_url": df[(df['App_Name'] == row['App_Name']) & (df['reviewer'] == row['reviewer'])].iloc[0].get('folder_url', '#')
        }

def build_dashboard():
    container = widgets.VBox()
    btn_export = widgets.Button(description="💾 Save Reports", button_style='success', icon='file-excel')
    btn_global = widgets.Button(description="📊 Global Report", button_style='primary')
    lbl_out = widgets.Label()

    app_widgets = []
    for app_key in sorted(unified_data.keys()):
        app_data = unified_data[app_key]
        pct = int((app_data['stats']['completed_users'] / app_data['stats']['total_users'] * 100)) if app_data['stats']['total_users'] > 0 else 0
        w_lbl = widgets.HTML(f"<b>📂 {app_key}</b> &nbsp; <span style='background:#eee; padding:2px 5px; border-radius:4px'>{pct}% Done</span>", layout=widgets.Layout(width='400px'))
        w_stat = widgets.Dropdown(options=["Calculated", "Force Completed", "Action Required"], value=app_data['status_manual'], layout=widgets.Layout(width='150px'))
        widget_store[app_key] = {"data": app_data, "w_stat": w_stat}

        rev_list = widgets.VBox([
            widgets.HBox([
                widgets.HTML(f"<a href='{rd['folder_url']}' target='_blank'>{rn}</a>", layout=widgets.Layout(width='250px')),
                widgets.HTML(rd['status_calc'], layout=widgets.Layout(width='200px')),
                widgets.HTML(rd['detail_html'])
            ]) for rn, rd in app_data['reviewers'].items()
        ], layout=widgets.Layout(margin='5px 0 5px 20px', display='none'))

        btn_tog = widgets.Button(description="➕ Show Users", layout=widgets.Layout(width='100px'))
        def create_tog(w):
            def on_tog(b):
                if w.layout.display == 'none': w.layout.display = 'block'; b.description = "➖ Hide"
                else: w.layout.display = 'none'; b.description = "➕ Show Users"
            return on_tog
        btn_tog.on_click(create_tog(rev_list))
        app_widgets.append(widgets.VBox([widgets.HBox([btn_tog, w_lbl, w_stat]), rev_list]))

    def export(b):
        b.disabled=True; b.description="Saving..."
        saved_files = []
        for app_key, widget_data in widget_store.items():
            app_name = widget_data['data']['App_Name']
            category = widget_data['data']['Category']
            app_rows = df[(df['Category'] == category) & (df['App_Name'] == app_name)].to_dict('records')

            final_data = []
            for row in app_rows:
                fin_st = widget_data['w_stat'].value
                if fin_st == "Calculated": fin_st = "Cached - Completed" if row.get('source_is_cache') else "Completed"

                final_data.append({
                    "User Name": row.get('user_name', ''),
                    "User Email": row.get('user_email', ''),
                    "Reviewer": row['reviewer'],
                    "File Name": row.get('file_name'),
                    "Reviewer Response": row.get('response'),
                    "Details of Access Change": row.get('details'),
                    "Final Status": fin_st,
                    "Row Num": row.get('row_number'),
                    "Audit Log": row.get('Audit_Log')
                })

            if final_data:
                safe_name = re.sub(r'[\\/*?:"<>|]', "", app_name)
                f_name = f"{output_dir}/{safe_name}_{int(time.time())}.xlsx"
                pd.DataFrame(final_data).to_excel(f_name, index=False)
                saved_files.append(safe_name)

        lbl_out.value = f"Saved {len(saved_files)} files."
        b.disabled=False; b.description="💾 Save Reports"

    def export_global(b):
        b.disabled = True; b.description = "Saving..."
        rows = []
        for _, r in stats.sort_values("App_Name").iterrows():
            rows.append({
                "Category": r['Category'],
                "App Name": r['App_Name'],
                "Reviewer": r['reviewer'],
                "Final Status": "Completed" if r['missing'] == 0 else "Pending",
                "Total Approved": int(r['approved']),
                "Total Denied": int(r['denied']),
                "Total Changed": int(r['changed'])
            })
        fname = f"{output_dir}/Global_Report_{int(time.time())}.xlsx"
        pd.DataFrame(rows).to_excel(fname, index=False)
        lbl_out.value = f"Global report saved."
        b.disabled = False; b.description = "📊 Global Report"

    btn_export.on_click(export)
    btn_global.on_click(export_global)
    container.children = tuple([widgets.HBox([btn_export, btn_global, lbl_out])] + app_widgets)
    display(container)

if unified_data:
    build_dashboard()
else:
    print("⚠️ No data found.")
