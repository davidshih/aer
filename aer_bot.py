import os
import logging
import requests
import pandas as pd
from datetime import datetime
from urllib.parse import quote
from dotenv import load_dotenv
from msal import PublicClientApplication
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

# === TUI 介面套件 ===
from rich.console import Console
from rich.table import Table
from rich.panel import Panel
from rich.logging import RichHandler
from InquirerPy import inquirer
from InquirerPy.base.control import Choice
from InquirerPy.separator import Separator

# === 初始化 ===
load_dotenv()
console = Console()

# 設定 Logging
logging.basicConfig(
    level="INFO",
    format="%(message)s",
    datefmt="[%X]",
    handlers=[RichHandler(console=console, rich_tracebacks=True)]
)
logger = logging.getLogger("aer")

# === 全域變數 ===
TENANT_ID = os.getenv("AZURE_TENANT_ID")
CLIENT_ID = os.getenv("AZURE_CLIENT_ID")
SHAREPOINT_HOST = os.getenv("SHAREPOINT_HOST", "davidshih.sharepoint.com")
SITE_NAME = os.getenv("SITE_NAME", "aer")
SENDER_EMAIL = os.getenv("SENDER_EMAIL") 
DEFAULT_DOMAIN = os.getenv("DEFAULT_DOMAIN", "company.com") # 預設網域

headers = {} # Token 容器

# ===========================
# 1. 核心 API 與 Helper 函數
# ===========================

def get_auth_token():
    """Device Code Flow 登入"""
    app = PublicClientApplication(
        CLIENT_ID,
        authority=f"https://login.microsoftonline.com/{TENANT_ID}"
    )
    scopes = ["Sites.Read.All", "Mail.Send", "User.Read.All"]
    flow = app.initiate_device_flow(scopes=scopes)
    
    if "user_code" not in flow:
        raise ValueError("無法建立 Device Flow")

    console.print(Panel(
        f"[bold yellow]請打開瀏覽器前往:[/bold yellow] {flow['verification_uri']}\n"
        f"[bold green]輸入代碼:[/bold green] {flow['user_code']}",
        title="🔐 需要登入", border_style="blue"
    ))
    
    result = app.acquire_token_by_device_flow(flow)
    if "access_token" in result:
        global headers
        headers = {"Authorization": f"Bearer {result['access_token']}"}
        console.print(f"[green]🎉 登入成功！[/green] User: {result.get('id_token_claims', {}).get('name')}")
        return True
    return False

def get_site_id(site_name):
    url = f"https://graph.microsoft.com/v1.0/sites/{SHAREPOINT_HOST}:/sites/{site_name}"
    resp = requests.get(url, headers=headers)
    resp.raise_for_status()
    return resp.json()["id"]

def list_folders(site_id, path):
    """列出資料夾 (支援 Root)"""
    if not path:
        url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/root/children"
    else:
        clean_path = path.strip("/")
        url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/root:/{clean_path}:/children"
    resp = requests.get(url, headers=headers)
    return [item for item in resp.json().get("value", []) if item.get("folder")]

def list_excel_files(site_id, folder_path):
    """列出資料夾下的 Excel"""
    clean_path = folder_path.strip("/")
    url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/root:/{clean_path}:/children"
    resp = requests.get(url, headers=headers)
    files = []
    for item in resp.json().get("value", []):
        if item["name"].endswith(".xlsx"):
            files.append({
                "id": item["id"],
                "name": item["name"],
                "lastModifiedDateTime": item.get("lastModifiedDateTime"),
                "webUrl": item.get("webUrl")
            })
    return sorted(files, key=lambda f: f.get("lastModifiedDateTime", ""), reverse=True)

def download_file(site_id, file_path):
    clean_path = file_path.strip("/")
    url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/root:/{clean_path}:/content"
    resp = requests.get(url, headers=headers)
    return resp.content

def get_file_audit_log(site_id, file_id):
    url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/items/{file_id}/versions"
    resp = requests.get(url, headers=headers)
    if resp.status_code != 200: return "無法取得版本紀錄"
    
    logs = []
    for v in resp.json().get("value", []):
        mod_time = v.get("lastModifiedDateTime", "")[:19].replace("T", " ")
        user_info = v.get("lastModifiedBy", {}).get("user", {})
        actor = user_info.get("displayName") or "Unknown"
        logs.append(f"{mod_time} - {actor}")
    return "\n".join(logs)

def read_visible_rows(excel_bytes, reviewer_name, file_name, folder_url):
    """讀取 Excel"""
    wb = load_workbook(BytesIO(excel_bytes))
    ws = wb.active
    
    # 簡單欄位對應
    header_row = [cell.value for cell in ws[1]]
    col_map = {str(name).strip(): idx for idx, name in enumerate(header_row) if name}
    
    COL_REVIEWER = "Reviewer"
    COL_RESPONSE = "Reviewer's Response"
    
    reviewer_col = col_map.get(COL_REVIEWER)
    response_col = col_map.get(COL_RESPONSE)
    
    if reviewer_col is None or response_col is None:
        return []

    results = []
    for row_idx in range(2, ws.max_row + 1):
        if ws.row_dimensions.get(row_idx) and ws.row_dimensions[row_idx].hidden:
            continue
            
        row = [cell.value for cell in ws[row_idx]]
        if not row: continue

        r_val = row[reviewer_col] if reviewer_col < len(row) else None
        resp_val = row[response_col] if response_col < len(row) else None
        
        if str(r_val).strip().lower() != reviewer_name.lower():
            continue
            
        is_missing = resp_val is None or str(resp_val).strip() == ""
        
        results.append({
            "reviewer": reviewer_name,
            "is_missing": is_missing,
            "response": resp_val,
            "folder_url": folder_url,
            "file_name": file_name
        })
    return results

def get_user_email(name, fallback_domain="company.com"):
    """
    查 Email (三階段策略)
    1. Graph API 精確搜尋
    2. Graph API 模糊搜尋
    3. 猜測法 (First.Last@domain.com)
    """
    try:
        clean_name = name.split("(")[0].strip()
        encoded = quote(clean_name)
        
        # 1. API 查詢
        url = f"https://graph.microsoft.com/v1.0/users?$filter=startswith(displayName, '{encoded}')"
        resp = requests.get(url, headers=headers)
        if resp.json().get("value"):
            u = resp.json()["value"][0]
            return u.get("mail") or u.get("userPrincipalName")
    except:
        pass
    
    # 2. 猜測法 (Fallback)
    try:
        parts = name.strip().split()
        if len(parts) >= 2:
            fname = parts[0]
            lname = parts[-1]
            return f"{fname}.{lname}@{fallback_domain}".lower()
        else:
            return f"{parts[0]}@{fallback_domain}".lower()
    except:
        return ""

def send_mail(sender, to, subject, body):
    """
    寄信 (支援 sender 參數)
    """
    # 如果 sender 是 "me" 或空，使用 API 預設路徑 /me/sendMail
    if not sender or sender.lower() == "me":
        url = "https://graph.microsoft.com/v1.0/me/sendMail"
    else:
        # 如果指定了別的信箱 (Shared Mailbox)，使用 /users/{sender}/sendMail
        url = f"https://graph.microsoft.com/v1.0/users/{sender}/sendMail"
        
    data = {
        "message": {
            "subject": subject,
            "body": {"contentType": "HTML", "content": body},
            "toRecipients": [{"emailAddress": {"address": to}}]
        }
    }
    resp = requests.post(url, headers={**headers, "Content-Type": "application/json"}, json=data)
    return resp.status_code == 202, resp.text

# ===========================
# 2. 報表生成 (Cell 7 & 8 Logic)
# ===========================

def format_excel_layout(file_path):
    """Excel 美化 (AutoFit + Wrap Text)"""
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        long_text_cols = ["Audit_History", "Details of Access change", "details", "response", "error"]
        
        for column_cells in ws.columns:
            header_cell = column_cells[0]
            header_val = str(header_cell.value).strip() if header_cell.value else ""
            col_letter = get_column_letter(header_cell.column)
            
            if header_val in long_text_cols:
                ws.column_dimensions[col_letter].width = 50
                for cell in column_cells:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
            else:
                max_length = 0
                for cell in column_cells:
                    try:
                        if cell.value:
                            cell_len = len(str(cell.value))
                            if cell_len > max_length: max_length = cell_len
                    except: pass
                
                adjusted = (max_length + 2) * 1.1
                if adjusted > 60: adjusted = 60
                ws.column_dimensions[col_letter].width = adjusted
                for cell in column_cells:
                    cell.alignment = Alignment(vertical='top')

        wb.save(file_path)
    except Exception as e:
        logger.warning(f"美化失敗: {e}")

def generate_reports(df, errors, app_name):
    """生成所有報表 (Cell 7 & 8)"""
    os.makedirs("output", exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_name = app_name.replace("/", "_")
    
    console.print(f"\n[bold cyan]📄 正在產生報表...[/bold cyan]")

    # 1. 成功資料報表
    if len(df) > 0:
        # Consolidated Report
        cons_file = f"output/consolidated_report_{safe_name}_{ts}.xlsx"
        df.to_excel(cons_file, index=False)
        format_excel_layout(cons_file)
        console.print(f"  ✅ 合併報告: [underline]{cons_file}[/underline]")
        
        # Missing Report
        missing_df = df[df["is_missing"]].copy()
        if len(missing_df) > 0:
            miss_file = f"output/missing_responses_{safe_name}_{ts}.xlsx"
            missing_df.to_excel(miss_file, index=False)
            format_excel_layout(miss_file)
            console.print(f"  ✅ 缺漏報告: [underline]{miss_file}[/underline]")

    # 2. 錯誤報告 (Cell 8)
    if errors:
        err_df = pd.DataFrame(errors)
        err_file = f"output/errors_{safe_name}_{ts}.xlsx"
        err_df.to_excel(err_file, index=False)
        format_excel_layout(err_file)
        console.print(f"  ⚠️ [bold red]錯誤報告:[/bold red] [underline]{err_file}[/underline]")
    else:
        console.print("  🎉 沒有錯誤，完美！")
    
    console.print(Panel("所有報表已生成完畢", border_style="green"))

# ===========================
# 3. 互動式介面 (CLI UI)
# ===========================

def folder_navigator(site_id):
    current_path = ""
    while True:
        console.clear()
        console.print(f"📂 目前路徑: [bold cyan]{current_path or 'Root'}[/bold cyan]")
        
        folders = list_folders(site_id, current_path)
        folder_names = sorted([f['name'] for f in folders if f['name'] != "Forms"])
        
        choices = []
        if current_path:
            choices.append(Choice(value="..", name="⬅️  返回上一層"))
        
        choices.append(Separator("--- 資料夾 ---"))
        for name in folder_names:
            choices.append(Choice(value=name, name=f"📂 {name}"))
            
        choices.append(Separator("--- 操作 ---"))
        choices.append(Choice(value="SELECT_THIS", name="✅ 選定此路徑為目標"))
        
        selection = inquirer.select(
            message="請選擇資料夾進入，或選定目標:",
            choices=choices,
            default=choices[0] if choices else None
        ).execute()
        
        if selection == "SELECT_THIS":
            return current_path
        elif selection == "..":
            current_path = "/".join(current_path.rstrip("/").split("/")[:-1])
        else:
            if current_path:
                current_path = f"{current_path}/{selection}"
            else:
                current_path = selection

def email_wizard(df):
    """
    升級版郵件精靈：循環選單 + 單獨編輯功能
    """
    missing_df = df[df['is_missing']].copy()
    if len(missing_df) == 0:
        return

    console.print(f"\n[bold red]⚠️ 發現 {len(missing_df)} 筆缺漏！準備進入郵件發送中心...[/bold red]")
    
    # === 設定初始資料 ===
    # 詢問網域
    fallback_domain = inquirer.text(message="請輸入公司網域 (用於自動猜測 Email):", default=DEFAULT_DOMAIN).execute()
    
    # 預設資料
    default_subject = "[Action Required] Access Review Reminder"
    default_sender = SENDER_EMAIL if SENDER_EMAIL else "me"
    default_body_tmpl = "<p>Hi {name},</p><p>請完成 <b>{app}</b> 的審核 (尚缺 {missing} 筆)。</p><p><a href='{link}'>連結</a></p>"
    
    # 準備資料結構
    user_data_map = {}
    unique_reviewers = missing_df.groupby(['App_Name', 'reviewer']).first().reset_index()
    
    with console.status("[bold green]正在建立寄信清單...[/bold green]"):
        for _, row in unique_reviewers.iterrows():
            name = row['reviewer']
            app = row.get('App_Name', 'App')
            # 使用新版 get_user_email (含 fallback)
            email = get_user_email(name, fallback_domain)
            missing_count = len(missing_df[(missing_df['reviewer'] == name) & (missing_df['App_Name'] == app)])
            
            key = f"{app}_{name}"
            user_data_map[key] = {
                "name": name,
                "app": app,
                "email": email, # To
                "sender": default_sender, # From (每個人可以不一樣)
                "missing": missing_count,
                "link": row['folder_url'],
                "selected": True # 預設全選
            }

    # === 主控制迴圈 (Control Loop) ===
    while True:
        # 統計選取狀態
        selected_users = [k for k, v in user_data_map.items() if v['selected']]
        
        console.clear()
        console.print(Panel(f"[bold white]📧 郵件發送控制台[/bold white]\n"
                            f"預設主旨: {default_subject}\n"
                            f"預設寄件: {default_sender}\n"
                            f"已選取人數: [bold cyan]{len(selected_users)} / {len(user_data_map)}[/bold cyan]",
                            title="Dashboard", border_style="cyan"))

        # 選單
        action = inquirer.select(
            message="請選擇操作:",
            choices=[
                Choice("SEND", f"🚀 發送郵件 ({len(selected_users)} 封)"),
                Choice("EDIT_USER", "✏️ 編輯個別使用者 (Email/Sender)"),
                Choice("TOGGLE", "✅ 勾選/取消發送對象"),
                Separator(),
                Choice("SET_SUBJ", "📝 修改全域主旨 (Global Subject)"),
                Choice("SET_BODY", "📝 修改內容模板 (Global Body)"),
                Choice("SET_FROM", "📧 修改預設寄件人 (Global Sender)"),
                Separator(),
                Choice("EXIT", "❌ 離開 (不發送)")
            ]
        ).execute()

        if action == "EXIT":
            console.print("再見！")
            break
            
        elif action == "SET_SUBJ":
            default_subject = inquirer.text(message="新主旨:", default=default_subject).execute()
            
        elif action == "SET_BODY":
            default_body_tmpl = inquirer.text(message="新內容 HTML:", default=default_body_tmpl).execute()
            
        elif action == "SET_FROM":
            default_sender = inquirer.text(message="新寄件人 (me 或 shared@com...):", default=default_sender).execute()
            # 更新所有人的 sender (除非之前有手動改過? 這裡簡單處理全改)
            for v in user_data_map.values():
                v['sender'] = default_sender

        elif action == "TOGGLE":
            # 使用 checkbox 讓使用者勾選
            choices = []
            for k, v in user_data_map.items():
                label = f"{v['app']} - {v['name']} ({v['email']})"
                choices.append(Choice(k, name=label, enabled=v['selected']))
            
            new_selection = inquirer.checkbox(
                message="請勾選要發送的對象:",
                choices=choices,
                cycle=False
            ).execute()
            
            # 更新選取狀態
            for k in user_data_map:
                user_data_map[k]['selected'] = (k in new_selection)

        elif action == "EDIT_USER":
            # 選擇要編輯的人
            user_choices = [Choice(k, f"{v['app']} - {v['name']}") for k, v in user_data_map.items()]
            target_key = inquirer.fuzzy(
                message="搜尋並選擇要編輯的使用者:",
                choices=user_choices,
            ).execute()
            
            if target_key:
                u = user_data_map[target_key]
                console.print(f"\n[bold]正在編輯: {u['name']}[/bold]")
                
                # 編輯各個欄位
                u['email'] = inquirer.text(message="收件人 (To):", default=u['email']).execute()
                u['sender'] = inquirer.text(message="寄件人 (From):", default=u['sender']).execute()
                # 這裡甚至可以讓使用者預覽這一封信
                console.print(f"[green]✅ 已更新 {u['name']} 的設定[/green]")
                inquirer.text(message="按 Enter 繼續...").execute()

        elif action == "SEND":
            if len(selected_users) == 0:
                console.print("[red]❌ 未選取任何對象！[/red]")
                inquirer.text(message="按 Enter 繼續...").execute()
                continue
                
            if not inquirer.confirm(message="確定要立即發送嗎?", default=True).execute():
                continue
            
            # 執行發送
            sent_count = 0
            with console.status("[bold yellow]🚀 飛鴿傳書中...[/bold yellow]"):
                for k in selected_users:
                    data = user_data_map[k]
                    
                    # 組合內容
                    final_body = default_body_tmpl.format(
                        name=data['name'],
                        app=data['app'],
                        missing=data['missing'],
                        link=data['link']
                    )
                    
                    success, msg = send_mail(data['sender'], data['email'], default_subject, final_body)
                    
                    if success:
                        console.print(f"  ✅ [green]Sent:[/green] {data['name']} ({data['email']})")
                        sent_count += 1
                        # 寄完自動取消選取，避免重複寄
                        user_data_map[k]['selected'] = False
                    else:
                        console.print(f"  ❌ [red]Fail:[/red] {data['name']} - {msg}")
            
            console.print(f"\n[bold cyan]發送作業結束。成功: {sent_count} 封。[/bold cyan]")
            inquirer.text(message="按 Enter 回到選單...").execute()

# ===========================
# 4. 主程式流程
# ===========================
def main():
    try:
        # Step 1: 登入
        if not get_auth_token(): return
        site_id = get_site_id(SITE_NAME)
        
        # Step 2: 選資料夾
        target_path = folder_navigator(site_id)
        
        # 接上變數
        BASE_PATH = target_path 
        APP_NAME = BASE_PATH.split("/")[-1]
        
        console.print(f"\n🎯 目標路徑: [bold yellow]{BASE_PATH}[/bold yellow]")
        
        # === Step 3: 掃描資料 (Cell 4 Logic) ===
        all_responses = []
        errors = [] # 收集錯誤
        
        initial_folders = list_folders(site_id, BASE_PATH)
        target_app_folders = []

        if initial_folders:
            first_folder_name = initial_folders[0]["name"]
            first_folder_path = f"{BASE_PATH}/{first_folder_name}"
            files_inside = list_excel_files(site_id, first_folder_path)
            is_user_folder = any(first_folder_name.lower() in f["name"].lower() for f in files_inside)
            
            if is_user_folder:
                target_app_folders.append((APP_NAME, BASE_PATH))
            else:
                for f in initial_folders:
                    target_app_folders.append((f['name'], f"{BASE_PATH}/{f['name']}"))
        else:
            logger.warning("目標路徑下是空的！")

        with console.status("[bold green]正在讀取 Excel...[/bold green]"):
            for current_app_name, current_path in target_app_folders:
                try:
                    reviewers = list_folders(site_id, current_path)
                    for folder in reviewers:
                        r_name = folder["name"]
                        f_url = folder.get("webUrl", "")
                        f_path = f"{current_path}/{r_name}"
                        
                        excel_files = list_excel_files(site_id, f_path)
                        target_files = [f for f in excel_files if r_name.lower() in f["name"].lower()]
                        
                        if target_files:
                            t_file = target_files[0]
                            # 抓取 Audit
                            audit_log = get_file_audit_log(site_id, t_file['id'])
                            content = download_file(site_id, f"{f_path}/{t_file['name']}")
                            rows = read_visible_rows(content, r_name, t_file['name'], f_url)
                            
                            for row in rows:
                                row["App_Name"] = current_app_name
                                row["Audit_History"] = audit_log
                                row["Last_Modified"] = t_file.get("lastModifiedDateTime")
                            
                            all_responses.extend(rows)
                        else:
                            # 找不到檔案不算錯誤，但可以 Log
                            pass
                except Exception as e:
                    errors.append({"reviewer": current_app_name, "error": str(e)})

        df = pd.DataFrame(all_responses)
        console.print(f"✅ 掃描完成！共發現 {len(df)} 筆資料。")
        
        # === Step 4: 自動產生報告 (Cell 7 & 8 Logic) ===
        # 這裡會自動把 Excel 報告生出來存到 output/
        generate_reports(df, errors, APP_NAME)
        
        if len(df) > 0:
            # === Step 5: 顯示狀態表 (Cell 5 Logic) ===
            table = Table(title="📊 審核狀態")
            table.add_column("App", style="cyan")
            table.add_column("Reviewer", style="magenta")
            table.add_column("Missing", justify="right")
            table.add_column("Status")
            
            summary = df.groupby(['App_Name', 'reviewer']).agg(
                missing_count=('is_missing', 'sum')
            ).reset_index()
            
            for _, row in summary.iterrows():
                miss = row['missing_count']
                status = "[red]未完成[/red]" if miss > 0 else "[green]完成[/green]"
                table.add_row(row['App_Name'], row['reviewer'], str(miss), status)
            
            console.print(table)
            
            # === Step 6: 寄信精靈 (Cell 6 Logic) ===
            email_wizard(df)
        
    except Exception as e:
        console.print_exception()

if __name__ == "__main__":
    main()