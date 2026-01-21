import sys
import os
import shutil
import time
import glob
import platform
import threading
import subprocess
from datetime import datetime

# GUI Imports
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext

# Logic Imports
try:
    import pandas as pd
    from openpyxl import load_workbook
except ImportError:
    print("Installing required packages...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "pandas", "openpyxl"])
    import pandas as pd
    from openpyxl import load_workbook

# Windows COM Check
WIN32COM_AVAILABLE = False
if platform.system() == 'Windows':
    try:
        import win32com.client
        import pythoncom
        WIN32COM_AVAILABLE = True
    except ImportError:
        pass

# ==========================================
# 1. Global Variables & Helper Functions
# ==========================================

excel_com_instance = None
log_file_handle = None

def sanitize_folder_name(name: str) -> str:
    invalid_chars = ['/', '\\', ':', '*', '?', '"', '<', '>', '|', '#', '%']
    sanitized = str(name).strip()
    for char in invalid_chars:
        sanitized = sanitized.replace(char, '_')
    return sanitized[:255].rstrip()

def find_column(worksheet, column_name):
    for col_idx, cell in enumerate(worksheet[1], start=1):
        if cell.value == column_name:
            return col_idx
    raise ValueError(f"Column '{column_name}' not found!")

def copy_selected_documents(source_dir, dest_dir, logger):
    """Automatically copies Word and PDF files."""
    # Copy Word
    for pattern in ["*.docx", "*.doc"]:
        for file in glob.glob(os.path.join(source_dir, pattern)):
            try:
                shutil.copy2(file, os.path.join(dest_dir, os.path.basename(file)))
                logger(f"  📎 Copied Word: {os.path.basename(file)}")
            except Exception as e:
                logger(f"  ⚠️ Failed to copy Word doc: {e}")

    # Copy PDF
    for file in glob.glob(os.path.join(source_dir, "*.pdf")):
        try:
            shutil.copy2(file, os.path.join(dest_dir, os.path.basename(file)))
            logger(f"  📎 Copied PDF: {os.path.basename(file)}")
        except Exception as e:
            logger(f"  ⚠️ Failed to copy PDF: {e}")

def copy_additional_files_list(file_paths: list, dest_dir: str, logger):
    """Copy specific user-selected files to the destination"""
    if not file_paths: return
    
    for path in file_paths:
        if os.path.exists(path) and os.path.isfile(path):
            try:
                shutil.copy2(path, os.path.join(dest_dir, os.path.basename(path)))
                logger(f"  📎 Copied Extra: {os.path.basename(path)}")
            except Exception as e:
                logger(f"  ❌ Copy Error ({os.path.basename(path)}): {e}")
        else:
            logger(f"  ⚠️ File not found: {path}")

# ==========================================
# 2. Excel Logic (Robust)
# ==========================================

def initialize_excel_com(logger):
    global excel_com_instance
    if WIN32COM_AVAILABLE and excel_com_instance is None:
        try:
            pythoncom.CoInitialize()
            excel_com_instance = win32com.client.Dispatch("Excel.Application")
            excel_com_instance.Visible = False
            excel_com_instance.DisplayAlerts = False
            excel_com_instance.ScreenUpdating = False
            return True
        except Exception as e:
            logger(f"  ❌ COM Init Failed: {e}")
            return False
    return excel_com_instance is not None

def cleanup_excel_com():
    global excel_com_instance
    if excel_com_instance:
        try:
            excel_com_instance.Quit()
        except: pass
        excel_com_instance = None

def process_hide_rows(file_path, reviewer, column_name, output_folder, logger):
    """Method A: Hides rows using OpenPyXL (Non-destructive, Safer without Excel)"""
    try:
        r_name = sanitize_folder_name(str(reviewer))
        r_folder = os.path.join(output_folder, r_name)
        os.makedirs(r_folder, exist_ok=True)
        
        base, ext = os.path.splitext(os.path.basename(file_path))
        new_path = os.path.join(r_folder, f"{base} - {r_name}{ext}")
        
        shutil.copy2(file_path, new_path)
        
        wb = load_workbook(new_path, keep_vba=True)
        ws = wb.active
        col_idx = find_column(ws, column_name)
        
        rows_to_hide = []
        for r in range(2, ws.max_row + 1):
            val = ws.cell(r, col_idx).value
            if str(val).strip() != str(reviewer).strip():
                rows_to_hide.append(r)
        
        for r in rows_to_hide:
            ws.row_dimensions[r].hidden = True
            
        wb.save(new_path)
        wb.close()
        logger(f"  ✅ Processed (Rows Hidden): {os.path.basename(new_path)}")
        return True, r_folder
    except Exception as e:
        logger(f"  ❌ OpenPyXL Error: {e}")
        return False, None

def process_delete_rows_com(file_path, reviewer, column_name, output_folder, logger):
    """Method B: Deletes rows using Windows Excel COM (Cleaner Result)"""
    if not WIN32COM_AVAILABLE: return False, None
    global excel_com_instance
    if not initialize_excel_com(logger): return False, None
    
    wb_source = None
    wb_dest = None
    
    try:
        r_name = sanitize_folder_name(str(reviewer))
        r_folder = os.path.join(output_folder, r_name)
        os.makedirs(r_folder, exist_ok=True)
        
        base, ext = os.path.splitext(os.path.basename(file_path))
        dst_path = os.path.join(r_folder, f"{base} - {r_name}.xlsx")
        
        abs_src = os.path.abspath(file_path)
        abs_dst = os.path.abspath(dst_path)
        
        wb_source = excel_com_instance.Workbooks.Open(abs_src, ReadOnly=True)
        wb_source.SaveCopyAs(abs_dst)
        wb_source.Close(False)
        
        wb_dest = excel_com_instance.Workbooks.Open(abs_dst)
        ws = wb_dest.Worksheets(1)
        
        if ws.AutoFilterMode: ws.AutoFilterMode = False

        last_cell = ws.Cells.SpecialCells(11) # xlCellTypeLastCell
        last_row = last_cell.Row
        last_col = last_cell.Column
        
        col_idx = None
        for col in range(1, last_col + 1):
            if str(ws.Cells(1, col).Value).strip() == str(column_name).strip():
                col_idx = col
                break
                
        if not col_idx:
            logger(f"  ❌ Column '{column_name}' not found.")
            wb_dest.Close(False)
            return False, None
            
        full_range = ws.Range(ws.Cells(1, 1), ws.Cells(last_row, last_col))
        full_range.AutoFilter(Field=col_idx, Criteria1=f"<>{reviewer}")
        
        try:
            data_to_delete = full_range.Offset(1, 0).Resize(last_row - 1, last_col)
            visible_cells = data_to_delete.SpecialCells(12) # xlCellTypeVisible
            visible_cells.EntireRow.Delete()
        except Exception as e:
            pass # No rows to delete

        ws.AutoFilterMode = False
        wb_dest.Save()
        wb_dest.Close()
        logger(f"  ✅ Processed (Rows Deleted): {os.path.basename(dst_path)}")
        return True, r_folder

    except Exception as e:
        logger(f"  ❌ COM Error: {e}")
        try: wb_source.Close(False); except: pass
        try: wb_dest.Close(False); except: pass
        return False, None

# ==========================================
# 3. GUI Application
# ==========================================

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Excel Reviewer Splitter Tool")
        self.geometry("720x800")
        
        # Variables
        self.file_path_var = tk.StringVar()
        self.col_name_var = tk.StringVar(value="Reviewer")
        self.out_dir_var = tk.StringVar()
        self.mode_var = tk.StringVar(value="delete" if WIN32COM_AVAILABLE else "hide")
        self.extra_files = [] # List to store file paths
        
        self.create_widgets()
        
    def create_widgets(self):
        # --- File & Folder Settings ---
        pnl = ttk.LabelFrame(self, text="File & Folder Settings", padding=10)
        pnl.pack(fill="x", padx=10, pady=5)
        
        # Excel Input
        ttk.Label(pnl, text="Excel File:").grid(row=0, column=0, sticky="w")
        ttk.Entry(pnl, textvariable=self.file_path_var, width=55).grid(row=0, column=1, padx=5)
        ttk.Button(pnl, text="Browse", command=self.browse_file).grid(row=0, column=2)
        
        # Column Name
        ttk.Label(pnl, text="Column Name:").grid(row=1, column=0, sticky="w", pady=5)
        ttk.Entry(pnl, textvariable=self.col_name_var, width=20).grid(row=1, column=1, sticky="w", padx=5)
        
        # Output Folder
        ttk.Label(pnl, text="Output Folder:").grid(row=2, column=0, sticky="w")
        ttk.Entry(pnl, textvariable=self.out_dir_var, width=55).grid(row=2, column=1, padx=5)
        ttk.Button(pnl, text="Browse", command=self.browse_folder).grid(row=2, column=2)

        # --- Additional Files ---
        pnl_files = ttk.LabelFrame(self, text="Additional Files to Attach (Multiple)", padding=10)
        pnl_files.pack(fill="x", padx=10, pady=5)
        
        btn_box = ttk.Frame(pnl_files)
        btn_box.pack(anchor="w", pady=(0, 5))
        ttk.Button(btn_box, text="➕ Add Files...", command=self.add_extra_files).pack(side="left", padx=5)
        ttk.Button(btn_box, text="🗑️ Clear List", command=self.clear_extra_files).pack(side="left")
        
        # Listbox with Scrollbar
        list_frame = ttk.Frame(pnl_files)
        list_frame.pack(fill="x")
        self.lst_files = tk.Listbox(list_frame, height=5, selectmode="extended", activestyle='none')
        self.lst_files.pack(side="left", fill="x", expand=True)
        
        scrollbar = ttk.Scrollbar(list_frame, orient="vertical", command=self.lst_files.yview)
        scrollbar.pack(side="right", fill="y")
        self.lst_files.config(yscrollcommand=scrollbar.set)

        # --- Processing Mode ---
        pnl_opt = ttk.LabelFrame(self, text="Processing Mode", padding=10)
        pnl_opt.pack(fill="x", padx=10, pady=5)
        
        ttk.Label(pnl_opt, text="How should the non-reviewer rows be handled?").pack(anchor="w")
        
        # Radio Buttons for Hide vs Delete
        
        # Option 1: Hide
        rb_hide = ttk.Radiobutton(pnl_opt, text="Hide Rows (Fast, preserves all data, just hidden)", 
                                  variable=self.mode_var, value="hide")
        rb_hide.pack(anchor="w", pady=2)
        
        # Option 2: Delete
        delete_text = "Delete Rows (Cleaner file, requires Windows Excel)"
        if not WIN32COM_AVAILABLE:
            delete_text += " [UNAVAILABLE - Excel not found]"
        
        rb_delete = ttk.Radiobutton(pnl_opt, text=delete_text, 
                                    variable=self.mode_var, value="delete",
                                    state="normal" if WIN32COM_AVAILABLE else "disabled")
        rb_delete.pack(anchor="w", pady=2)

        ttk.Label(pnl_opt, text="ℹ️ Word (.doc/x) and PDF (.pdf) files are always copied automatically.", foreground="gray").pack(anchor="w", pady=(5,0))

        # --- Run & Progress ---
        btn_frame = ttk.Frame(self)
        btn_frame.pack(pady=10)
        
        self.btn_run = ttk.Button(btn_frame, text="🚀 Start Processing", command=self.start_thread)
        self.btn_run.pack(side="left", padx=5)

        # Progress Label and Bar
        progress_frame = ttk.Frame(self)
        progress_frame.pack(fill="x", padx=10, pady=5)
        
        self.lbl_progress_text = ttk.Label(progress_frame, text="Ready", font=("Arial", 9, "bold"))
        self.lbl_progress_text.pack(anchor="w", pady=(0, 2))
        
        self.progress = ttk.Progressbar(progress_frame, orient="horizontal", mode="determinate")
        self.progress.pack(fill="x")

        # --- Log ---
        ttk.Label(self, text="Execution Log:").pack(anchor="w", padx=10)
        self.log_area = scrolledtext.ScrolledText(self, height=10, state='disabled')
        self.log_area.pack(fill="both", expand=True, padx=10, pady=5)

    def browse_file(self):
        f = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx *.xlsm")])
        if f: 
            self.file_path_var.set(f)
            # Automatically set output directory to the same folder as input
            parent_dir = os.path.dirname(f)
            self.out_dir_var.set(parent_dir)

    def browse_folder(self):
        d = filedialog.askdirectory()
        if d: self.out_dir_var.set(d)

    def add_extra_files(self):
        files = filedialog.askopenfilenames(title="Select additional files to attach")
        for f in files:
            if f not in self.extra_files:
                self.extra_files.append(f)
                self.lst_files.insert(tk.END, f)
    
    def clear_extra_files(self):
        self.extra_files = []
        self.lst_files.delete(0, tk.END)

    def log(self, msg, level="INFO"):
        ts = datetime.now().strftime("%H:%M:%S")
        full_msg = f"[{ts}] {msg}"
        self.log_area.config(state='normal')
        self.log_area.insert(tk.END, full_msg + "\n")
        self.log_area.see(tk.END)
        self.log_area.config(state='disabled')
        if self.log_file_handle:
            try:
                self.log_file_handle.write(full_msg + "\n")
                self.log_file_handle.flush()
            except: pass

    def start_thread(self):
        self.btn_run.config(state="disabled")
        t = threading.Thread(target=self.run_process)
        t.start()

    def run_process(self):
        file_path = self.file_path_var.get()
        col_name = self.col_name_var.get()
        out_folder = self.out_dir_var.get()
        mode = self.mode_var.get()
        
        # Create Log File in subfolder
        today_str = datetime.now().strftime("%Y-%m-%d")
        log_dir = os.path.join(out_folder, "log", today_str)
        try:
            os.makedirs(log_dir, exist_ok=True)
            log_path = os.path.join(log_dir, f"aer-share-{today_str}.log")
            self.log_file_handle = open(log_path, "a", encoding="utf-8")
        except:
            self.log("⚠️ Could not create log file path", "WARN")
        
        self.log(f"🚀 Starting Task (Mode: {mode.upper()})...")
        self.lbl_progress_text.config(text="Initializing...")
        
        try:
            if not os.path.exists(file_path):
                self.log("❌ Excel file not found!", "ERROR")
                return

            df = pd.read_excel(file_path, engine='openpyxl')
            if col_name not in df.columns:
                self.log(f"❌ Column '{col_name}' not found.", "ERROR")
                return

            reviewers = df[col_name].dropna().unique().tolist()
            total = len(reviewers)
            self.log(f"🔎 Found {total} reviewers.")
            
            self.progress["maximum"] = total
            self.progress["value"] = 0

            # Initialize Excel COM if we are in delete mode
            if mode == "delete" and WIN32COM_AVAILABLE:
                pythoncom.CoInitialize()

            for i, reviewer in enumerate(reviewers):
                # Update Text Label with Count
                display_text = f"Processing: {reviewer} ({i+1}/{total})"
                self.lbl_progress_text.config(text=display_text)
                self.update_idletasks()
                
                self.log(f"Processing ({i+1}/{total}): {reviewer}")
                
                success = False
                r_folder = None
                
                # BRANCH: Choose method based on Radio Button
                if mode == "delete" and WIN32COM_AVAILABLE:
                    success, r_folder = process_delete_rows_com(file_path, reviewer, col_name, out_folder, self.log)
                else:
                    success, r_folder = process_hide_rows(file_path, reviewer, col_name, out_folder, self.log)

                if success and r_folder:
                    # 1. Auto Copy Word/PDF from source dir
                    base_dir = os.path.dirname(file_path)
                    copy_selected_documents(base_dir, r_folder, self.log)
                    
                    # 2. Copy Extra selected files
                    copy_additional_files_list(self.extra_files, r_folder, self.log)
                
                self.progress["value"] = i + 1
            
            self.lbl_progress_text.config(text=f"Done! Processed {total} reviewers.")
            self.log("🎉 All tasks completed!")
            messagebox.showinfo("Done", "Processing Complete!")

        except Exception as e:
            self.log(f"❌ Critical Error: {e}")
            messagebox.showerror("Error", str(e))
        finally:
            if self.log_file_handle:
                self.log_file_handle.close()
            cleanup_excel_com()
            self.btn_run.config(state="normal")

if __name__ == "__main__":
    app = App()
    app.mainloop()