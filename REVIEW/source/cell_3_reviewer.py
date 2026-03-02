# === CELL 3: Stage 3 — Reviewer Assignment & Manual Override ===
# Assigns reviewers based on mapping file, outputs review Excel with dropdowns.

# ============================================
# STATE
# ============================================

s3_input_df = None
s3_mapping_df = None
s3_output_path = None
s3_status = widgets.HTML(value="<i>Ready. Upload validated file and mapping.</i>")
s3_output = widgets.Output()

s3_upload_validated = widgets.FileUpload(accept=".csv,.xlsx", multiple=False, description="Validated File")
s3_upload_mapping = widgets.FileUpload(accept=".csv,.xlsx", multiple=False, description="Mapping File")

s3_btn_assign = widgets.Button(description="⚙️ Assign Reviewers", button_style="success",
                                layout=widgets.Layout(width="200px", height="40px"), disabled=True)
s3_btn_save = widgets.Button(description="💾 Save Review File", button_style="info",
                              layout=widgets.Layout(width="200px", height="40px"), disabled=True)

# ============================================
# HELPERS
# ============================================

def _load_mapping_file(content, fname):
    """Load mapping file, supports both CSV and XLSX."""
    if fname.lower().endswith(".xlsx"):
        return pd.read_excel(io.BytesIO(content))
    encoding = "utf-8"
    if CHARDET_AVAILABLE:
        det = chardet.detect(content)
        encoding = det.get("encoding", "utf-8") or "utf-8"
    return pd.read_csv(io.BytesIO(content), encoding=encoding)


def _detect_mapping_cols(df):
    """Detect department and reviewer columns in mapping file."""
    dept_col = None
    reviewer_col = None
    email_col = None
    for col in df.columns:
        cl = str(col).strip().lower()
        if "department" in cl or "dept" in cl:
            dept_col = col
        if "reviewer" in cl or "head" in cl:
            reviewer_col = col
        if "email" in cl and ("reviewer" in cl or "head" in cl):
            email_col = col
    return dept_col, reviewer_col, email_col


# ============================================
# HANDLERS
# ============================================

def on_s3_upload_change(change):
    global s3_input_df, s3_mapping_df
    try:
        if s3_upload_validated.value:
            uploaded = list(s3_upload_validated.value.values())[0] if isinstance(s3_upload_validated.value, dict) else s3_upload_validated.value[0]
            content = uploaded["content"] if isinstance(uploaded, dict) else uploaded.content
            fname = uploaded["name"] if isinstance(uploaded, dict) else uploaded.name
            if fname.lower().endswith(".csv"):
                s3_input_df = pd.read_csv(io.BytesIO(content))
            else:
                s3_input_df = pd.read_excel(io.BytesIO(content))

        if s3_upload_mapping.value:
            uploaded = list(s3_upload_mapping.value.values())[0] if isinstance(s3_upload_mapping.value, dict) else s3_upload_mapping.value[0]
            content = uploaded["content"] if isinstance(uploaded, dict) else uploaded.content
            fname = uploaded["name"] if isinstance(uploaded, dict) else uploaded.name
            s3_mapping_df = _load_mapping_file(content, fname)

        if s3_input_df is not None and s3_mapping_df is not None:
            s3_btn_assign.disabled = False
            s3_status.value = f"<span style='color:green'>✅ Files loaded: {len(s3_input_df)} users, {len(s3_mapping_df)} mapping rows</span>"
        elif s3_input_df is not None:
            # Try auto-load latest mapping
            auto_mapping, auto_path = load_latest_mapping()
            if auto_mapping is not None:
                s3_mapping_df = auto_mapping
                s3_btn_assign.disabled = False
                s3_status.value = f"<span style='color:green'>✅ Validated: {len(s3_input_df)} rows | Auto-loaded mapping: {os.path.basename(auto_path)}</span>"
            else:
                s3_status.value = f"<span style='color:orange'>⚠️ Validated file loaded. Upload mapping or save one from Stage 1.5.</span>"
    except Exception as e:
        s3_status.value = f"<span style='color:red'>❌ {e}</span>"

s3_upload_validated.observe(on_s3_upload_change, names="value")
s3_upload_mapping.observe(on_s3_upload_change, names="value")


def on_s3_assign(_):
    global s3_output_path
    s3_output.clear_output()
    s3_btn_assign.disabled = True
    try:
        if s3_input_df is None:
            s3_status.value = "<span style='color:red'>❌ No validated file loaded</span>"
            return
        if s3_mapping_df is None:
            s3_status.value = "<span style='color:red'>❌ No mapping file loaded</span>"
            return

        dept_col, reviewer_col, email_col = _detect_mapping_cols(s3_mapping_df)
        if not dept_col or not (reviewer_col or email_col):
            s3_status.value = "<span style='color:red'>❌ Cannot detect department/reviewer columns in mapping file</span>"
            return

        # Build lookup: department -> reviewer email/name
        dept_to_reviewer = {}
        email_to_reviewer = {}
        for _, mrow in s3_mapping_df.iterrows():
            dept = str(mrow.get(dept_col, "")).strip().lower()
            rev = str(mrow.get(email_col or reviewer_col, "")).strip()
            if dept:
                dept_to_reviewer[dept] = rev
            rev_email = str(mrow.get(email_col, "")).strip().lower() if email_col else ""
            if rev_email and rev_email != "nan":
                email_to_reviewer[rev_email] = rev

        # Detect email/department columns in validated file
        vd_email_col = _detect_email_col(s3_input_df)
        vd_dept_col = None
        for col in s3_input_df.columns:
            cl = str(col).strip().lower()
            if "department" in cl or "dept" in cl:
                vd_dept_col = col
                break

        assigned = []
        for idx, row in s3_input_df.iterrows():
            row_dict = row.to_dict()
            user_email = str(row.get(vd_email_col, "")).strip().lower() if vd_email_col else ""
            user_dept = str(row.get(vd_dept_col, "")).strip().lower() if vd_dept_col else ""

            # Priority 1: exact email match in mapping (dept head)
            if user_email in email_to_reviewer:
                row_dict["Reviewer"] = email_to_reviewer[user_email]
                row_dict["dept_head"] = "yes"
            # Priority 2: department match
            elif user_dept:
                matched_rev = None
                for map_dept, rev in dept_to_reviewer.items():
                    if map_dept in user_dept or user_dept in map_dept:
                        matched_rev = rev
                        break
                row_dict["Reviewer"] = matched_rev or ""
                row_dict["dept_head"] = ""
            else:
                row_dict["Reviewer"] = ""
                row_dict["dept_head"] = ""

            row_dict["Reviewer's Response"] = ""
            row_dict["Details of Access Change"] = ""
            assigned.append(row_dict)

        out_df = pd.DataFrame(assigned)

        # Sort: dept heads first, then by reviewer
        out_df["_sort_key"] = out_df["dept_head"].apply(lambda x: 0 if x == "yes" else 1)
        out_df = out_df.sort_values(["_sort_key", "Reviewer"]).drop(columns=["_sort_key"])

        ts = datetime.now().strftime("%Y%m%d_%H%M")
        fname = f"review_{ts}.xlsx"
        s3_output_path = safe_excel_path(os.path.join(STAGE3_DIR, fname))
        out_df.to_excel(s3_output_path, index=False)

        # Add data validation dropdowns
        if OPENPYXL_AVAILABLE:
            from openpyxl.worksheet.datavalidation import DataValidation
            wb = load_workbook(s3_output_path)
            ws = wb.active
            # Find response column
            resp_col = None
            for col_idx in range(1, ws.max_column + 1):
                if ws.cell(row=1, column=col_idx).value == "Reviewer's Response":
                    resp_col = col_idx
                    break
            if resp_col:
                dv = DataValidation(
                    type="list",
                    formula1='"Approved,Denied,Changes Required"',
                    allow_blank=True
                )
                dv.error = "Please select a valid response"
                dv.errorTitle = "Invalid Response"
                ws.add_data_validation(dv)
                for row_idx in range(2, ws.max_row + 1):
                    dv.add(ws.cell(row=row_idx, column=resp_col))
            wb.save(s3_output_path)
            format_export_excel(s3_output_path)

        with s3_output:
            assigned_count = sum(1 for r in assigned if r.get("Reviewer"))
            print(f"Assigned {assigned_count}/{len(assigned)} users to reviewers")
            print(f"Dept heads: {sum(1 for r in assigned if r.get('dept_head') == 'yes')}")

        s3_status.value = f"<span style='color:green'>✅ Review file saved: {os.path.basename(s3_output_path)}</span>"
        s3_btn_save.disabled = False
        logger(f"Stage 3: {assigned_count}/{len(assigned)} users assigned, saved to {s3_output_path}")
    except Exception as e:
        s3_status.value = f"<span style='color:red'>❌ {e}</span>"
        logger(f"Stage 3 error: {e}", "error")
    finally:
        s3_btn_assign.disabled = False

s3_btn_assign.on_click(on_s3_assign)
s3_btn_save.on_click(lambda _: None)  # Already saved in assign step

# ============================================
# UI
# ============================================

stage3_ui = widgets.VBox([
    widgets.HTML("""
        <div style='background: linear-gradient(135deg, #fa709a 0%, #fee140 100%);
            padding: 20px; border-radius: 8px; color: white; margin-bottom: 20px;
            box-shadow: 0 4px 6px rgba(0,0,0,0.1);'>
            <h2 style='margin: 0 0 10px 0;'>⚙️ Stage 3: Reviewer Assignment</h2>
            <p style='margin: 0; opacity: 0.95;'>
                Assign reviewers based on mapping file (CSV or XLSX). Outputs review Excel with dropdowns.
            </p>
        </div>
    """),
    widgets.HBox([s3_upload_validated, s3_upload_mapping]),
    widgets.HBox([s3_btn_assign, s3_btn_save]),
    s3_status,
    s3_output,
])

clear_output()
display(stage3_ui)
