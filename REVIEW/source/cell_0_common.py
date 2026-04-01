# === CELL 0: Common Library & Configuration ===
# Shared utilities for ALL subsequent cells.
# Run this cell FIRST before any stage.

import os
import sys
import io
import re
import glob
import json
import time
import shutil
import base64
import html
import logging
import platform
import hashlib
from dataclasses import dataclass
from datetime import datetime, timedelta
from urllib.parse import quote, urlparse, unquote
from enum import Enum

import pandas as pd
import requests
import ipywidgets as widgets
from dotenv import load_dotenv
from msal import PublicClientApplication
from IPython.display import display, clear_output, HTML as IPHTML

try:
    from rapidfuzz import fuzz as rf_fuzz
    FUZZY_AVAILABLE = True
except ImportError:
    try:
        from fuzzywuzzy import fuzz as rf_fuzz
        FUZZY_AVAILABLE = True
    except ImportError:
        FUZZY_AVAILABLE = False
        rf_fuzz = None

try:
    import openpyxl
    from openpyxl.styles import Alignment, PatternFill, Font
    from openpyxl.utils import get_column_letter
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import chardet
    CHARDET_AVAILABLE = True
except ImportError:
    CHARDET_AVAILABLE = False

WIN32COM_AVAILABLE = False
if platform.system() == "Windows":
    try:
        import win32com.client
        import pythoncom
        WIN32COM_AVAILABLE = True
    except ImportError:
        pass

try:
    import tkinter as tk
    from tkinter import filedialog
    TK_AVAILABLE = True
except ImportError:
    TK_AVAILABLE = False


# ============================================
# RUNTIME CONFIGURATION
# ============================================

@dataclass(frozen=True)
class AppConfig:
    tenant_id: str
    client_id: str
    sharepoint_host: str
    site_name: str
    sender_email: str
    review_year: str
    root_person: str
    org_depth: int
    fuzzy_threshold: int
    base_path: str
    email_template_footer: str


@dataclass(frozen=True)
class RuntimePaths:
    today_str: str
    hour_str: str
    input_dir: str
    input_ad_cache_dir: str
    input_mapping_dir: str
    output_base_dir: str
    output_log_dir: str
    output_ad_cache_dir: str
    output_orgchart_dir: str
    stage2_dir: str
    stage3_dir: str
    stage4_dir: str
    report_dir: str
    cache_dir: str
    checkpoint_dir: str
    log_file: str


@dataclass(frozen=True)
class AppRuntime:
    config: AppConfig
    paths: RuntimePaths
    logger: logging.Logger
    token_manager_cls: type


_DEFAULT_NOW = datetime.now()
_DEFAULT_TODAY_STR = _DEFAULT_NOW.strftime("%Y-%m-%d")
_DEFAULT_HOUR_STR = _DEFAULT_NOW.strftime("%H")

INPUT_DIR = "input"
INPUT_AD_CACHE_DIR = os.path.join(INPUT_DIR, "ad_cache")
INPUT_MAPPING_DIR = os.path.join(INPUT_DIR, "mapping")
OUTPUT_BASE_DIR = os.path.join("output", _DEFAULT_TODAY_STR)
OUTPUT_LOG_DIR = os.path.join(OUTPUT_BASE_DIR, "logs")
OUTPUT_AD_CACHE_DIR = os.path.join(OUTPUT_BASE_DIR, "ad_cache")
OUTPUT_ORGCHART_DIR = os.path.join(OUTPUT_BASE_DIR, "orgchart")
STAGE2_DIR = os.path.join(OUTPUT_BASE_DIR, "stage2_validated")
STAGE3_DIR = os.path.join(OUTPUT_BASE_DIR, "stage3_review")
STAGE4_DIR = os.path.join(OUTPUT_BASE_DIR, "stage4_splitter")
REPORT_DIR = os.path.join(OUTPUT_BASE_DIR, "report")
CACHE_DIR = os.path.join(OUTPUT_BASE_DIR, "cache")
CHECKPOINT_DIR = os.path.join(OUTPUT_BASE_DIR, "checkpoints")

TENANT_ID = ""
CLIENT_ID = ""
SHAREPOINT_HOST = ""
SITE_NAME = "aer"
SENDER_EMAIL = ""
AER_REVIEW_YEAR = str(_DEFAULT_NOW.year)
AER_ROOT_PERSON = "Steven Bush"
AER_ORG_DEPTH = 3
AER_FUZZY_THRESHOLD = 89
BASE_PATH = f"{AER_REVIEW_YEAR} Entitlement Review"
AER_EMAIL_TEMPLATE_FOOTER = "Sincerely,\nApple Bank's Information Security Team"
TODAY_STR = _DEFAULT_TODAY_STR
HOUR_STR = _DEFAULT_HOUR_STR
LOG_FILE = os.path.join(OUTPUT_LOG_DIR, f"aer_{TODAY_STR}_{HOUR_STR}00.log")

EMAIL_PATTERN = re.compile(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[A-Za-z]{2,}$')
BRACKET_CHARS = set("[](){}<>")
KNOWN_DOMAIN_CORRECTIONS = {"apple-bank.com": "applebank.com"}

_RUNTIME = None


def _read_int_env(name, default):
    raw = os.getenv(name, str(default)).strip()
    try:
        return int(raw)
    except ValueError as exc:
        raise ValueError(f"{name} must be an integer, got {raw!r}") from exc


def load_config():
    """Load and validate environment configuration."""
    load_dotenv(override=True)

    review_year = os.getenv("AER_REVIEW_YEAR", str(datetime.now().year)).strip()
    base_path = os.getenv("BASE_PATH", "").strip()
    if not base_path:
        base_path = os.getenv("APP_NAME", f"{review_year} Entitlement Review").strip()

    config = AppConfig(
        tenant_id=os.getenv("AZURE_TENANT_ID", "").strip(),
        client_id=os.getenv("AZURE_CLIENT_ID", "").strip(),
        sharepoint_host=os.getenv("SHAREPOINT_HOST", "").strip(),
        site_name=os.getenv("SITE_NAME", "aer").strip(),
        sender_email=os.getenv("SENDER_EMAIL", "").strip(),
        review_year=review_year,
        root_person=os.getenv("AER_ROOT_PERSON", "Steven Bush").strip(),
        org_depth=_read_int_env("AER_ORG_DEPTH", 3),
        fuzzy_threshold=_read_int_env("AER_FUZZY_THRESHOLD", 89),
        base_path=base_path,
        email_template_footer=os.getenv(
            "AER_EMAIL_TEMPLATE_FOOTER",
            "Sincerely,\nApple Bank's Information Security Team",
        ).strip().replace("\\n", "\n"),
    )

    missing = []
    if not config.tenant_id:
        missing.append("AZURE_TENANT_ID")
    if not config.client_id:
        missing.append("AZURE_CLIENT_ID")
    if not config.sharepoint_host:
        missing.append("SHAREPOINT_HOST")
    if not config.sender_email:
        missing.append("SENDER_EMAIL")
    if missing:
        raise RuntimeError(f"Missing required environment settings: {', '.join(missing)}")

    return config


def _build_paths(config, now=None):
    runtime_now = now or datetime.now()
    today_str = runtime_now.strftime("%Y-%m-%d")
    hour_str = runtime_now.strftime("%H")
    output_base_dir = os.path.join("output", today_str)
    output_log_dir = os.path.join(output_base_dir, "logs")
    return RuntimePaths(
        today_str=today_str,
        hour_str=hour_str,
        input_dir=INPUT_DIR,
        input_ad_cache_dir=os.path.join(INPUT_DIR, "ad_cache"),
        input_mapping_dir=os.path.join(INPUT_DIR, "mapping"),
        output_base_dir=output_base_dir,
        output_log_dir=output_log_dir,
        output_ad_cache_dir=os.path.join(output_base_dir, "ad_cache"),
        output_orgchart_dir=os.path.join(output_base_dir, "orgchart"),
        stage2_dir=os.path.join(output_base_dir, "stage2_validated"),
        stage3_dir=os.path.join(output_base_dir, "stage3_review"),
        stage4_dir=os.path.join(output_base_dir, "stage4_splitter"),
        report_dir=os.path.join(output_base_dir, "report"),
        cache_dir=os.path.join(output_base_dir, "cache"),
        checkpoint_dir=os.path.join(output_base_dir, "checkpoints"),
        log_file=os.path.join(output_log_dir, f"aer_{today_str}_{hour_str}00.log"),
    )


def ensure_runtime_dirs(config):
    """Create the runtime directory tree and return resolved paths."""
    paths = _build_paths(config)
    for path in [
        paths.input_ad_cache_dir,
        paths.input_mapping_dir,
        paths.output_log_dir,
        paths.output_ad_cache_dir,
        paths.output_orgchart_dir,
        paths.stage2_dir,
        paths.stage3_dir,
        paths.stage4_dir,
        paths.report_dir,
        paths.cache_dir,
        paths.checkpoint_dir,
    ]:
        os.makedirs(path, exist_ok=True)
    return paths


# ============================================
# UNIFIED LOGGER
# ============================================

aer_logger = logging.getLogger("aer")
if not aer_logger.handlers:
    aer_logger.addHandler(logging.NullHandler())

_log_formatter = logging.Formatter("%(asctime)s | %(levelname)s | %(message)s")

def _get_console_stream():
    s = getattr(sys, "stdout", None)
    try:
        if s and hasattr(s, "reconfigure"):
            s.reconfigure(encoding="utf-8")
        if hasattr(sys, "stderr") and hasattr(sys.stderr, "reconfigure"):
            sys.stderr.reconfigure(encoding="utf-8")
    except Exception:
        pass
    return s or sys.stdout

def setup_logger(paths):
    """Configure the shared logger after runtime paths exist."""
    global aer_logger
    logger_obj = logging.getLogger("aer")
    logger_obj.handlers.clear()
    logger_obj.setLevel(logging.INFO)

    console_handler = logging.StreamHandler(_get_console_stream())
    console_handler.setFormatter(_log_formatter)
    logger_obj.addHandler(console_handler)

    file_handler = logging.FileHandler(paths.log_file, encoding="utf-8", mode="a")
    file_handler.setFormatter(_log_formatter)
    logger_obj.addHandler(file_handler)
    aer_logger = logger_obj
    return logger_obj

def logger(msg, level="info"):
    getattr(aer_logger, level, aer_logger.info)(msg)


# ============================================
# RATE-LIMITED HTTP CLIENT
# ============================================

HTTP_MAX_RETRIES = 3
HTTP_BACKOFF_BASE = 2.0
HTTP_TIMEOUT = 30

def _http_request(method, url, headers=None, json_payload=None, data=None, timeout=HTTP_TIMEOUT):
    """HTTP request with exponential backoff and 429 retry."""
    last_exc = None
    for attempt in range(HTTP_MAX_RETRIES + 1):
        try:
            resp = requests.request(
                method, url, headers=headers, json=json_payload, data=data, timeout=timeout
            )
            if resp.status_code == 429:
                retry_after = int(resp.headers.get("Retry-After", HTTP_BACKOFF_BASE ** (attempt + 1)))
                retry_after = min(retry_after, 60)
                logger(f"  [429] Rate limited, retrying in {retry_after}s (attempt {attempt+1})")
                time.sleep(retry_after)
                continue
            if resp.status_code >= 500 and attempt < HTTP_MAX_RETRIES:
                wait = HTTP_BACKOFF_BASE ** (attempt + 1)
                logger(f"  [5xx] Server error {resp.status_code}, retrying in {wait:.0f}s")
                time.sleep(wait)
                continue
            return resp
        except requests.exceptions.RequestException as e:
            last_exc = e
            if attempt < HTTP_MAX_RETRIES:
                wait = HTTP_BACKOFF_BASE ** (attempt + 1)
                logger(f"  [NET] {e}, retrying in {wait:.0f}s")
                time.sleep(wait)
    if last_exc:
        raise last_exc
    raise RuntimeError("HTTP request failed after retries")

def graph_get(url, headers, timeout=HTTP_TIMEOUT):
    return _http_request("GET", url, headers=headers, timeout=timeout)

def graph_post(url, headers, json_payload=None, timeout=HTTP_TIMEOUT):
    h = dict(headers or {})
    if json_payload is not None:
        h["Content-Type"] = "application/json"
    return _http_request("POST", url, headers=h, json_payload=json_payload, timeout=timeout)

def graph_put(url, headers, data=None, timeout=HTTP_TIMEOUT):
    return _http_request("PUT", url, headers=headers, data=data, timeout=timeout)


# ============================================
# TOKEN MANAGER
# ============================================

class TokenManager:
    """Manages OAuth tokens with auto-refresh."""

    def __init__(self, tenant_id, client_id):
        self.tenant_id = tenant_id
        self.client_id = client_id
        self.app = PublicClientApplication(
            client_id, authority=f"https://login.microsoftonline.com/{tenant_id}"
        ) if tenant_id and client_id else None
        self._tokens = {}  # scope_key -> {access_token, exp_ts}

    def _decode_exp(self, token):
        try:
            parts = token.split(".")
            if len(parts) < 2:
                return time.time() + 3600
            payload_raw = parts[1] + ("=" * (-len(parts[1]) % 4))
            payload = json.loads(base64.urlsafe_b64decode(payload_raw.encode("utf-8")))
            return payload.get("exp", time.time() + 3600)
        except Exception:
            return time.time() + 3600

    def _is_expired(self, scope_key, margin_sec=300):
        entry = self._tokens.get(scope_key)
        if not entry:
            return True
        return time.time() >= (entry["exp_ts"] - margin_sec)

    def login_interactive(self, scopes, scope_key="graph"):
        if not self.app:
            raise RuntimeError("Missing AZURE_TENANT_ID or AZURE_CLIENT_ID")
        result = self.app.acquire_token_interactive(scopes=scopes, prompt="select_account")
        if "access_token" not in result:
            err = result.get("error_description", "Unknown error")
            raise RuntimeError(f"Login failed: {err}")
        token = result["access_token"]
        self._tokens[scope_key] = {
            "access_token": token,
            "exp_ts": self._decode_exp(token),
            "scopes": scopes,
        }
        logger(f"✅ Login successful (scope_key={scope_key})")
        return token

    def _try_silent_refresh(self, scope_key):
        entry = self._tokens.get(scope_key)
        if not entry or not self.app:
            return False
        accounts = self.app.get_accounts()
        if not accounts:
            return False
        result = self.app.acquire_token_silent(scopes=entry["scopes"], account=accounts[0])
        if result and "access_token" in result:
            token = result["access_token"]
            self._tokens[scope_key] = {
                "access_token": token,
                "exp_ts": self._decode_exp(token),
                "scopes": entry["scopes"],
            }
            logger(f"🔄 Token refreshed (scope_key={scope_key})")
            return True
        return False

    def get_headers(self, scope_key="graph"):
        if self._is_expired(scope_key):
            if not self._try_silent_refresh(scope_key):
                raise RuntimeError(
                    f"Token expired for '{scope_key}'. Please re-run login cell."
                )
        token = self._tokens[scope_key]["access_token"]
        return {"Authorization": f"Bearer {token}"}

    def get_token(self, scope_key="graph"):
        headers = self.get_headers(scope_key)
        return headers["Authorization"].split(" ", 1)[1]

    def set_token(self, scope_key, token, scopes=None):
        self._tokens[scope_key] = {
            "access_token": token,
            "exp_ts": self._decode_exp(token),
            "scopes": scopes or [],
        }

    def has_token(self, scope_key="graph"):
        return scope_key in self._tokens


token_mgr = None


# ============================================
# AD CACHE HELPERS
# ============================================

def save_ad_cache(df, prefix="ad_users"):
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    fname = f"{prefix}_{ts}.csv"
    for d in [INPUT_AD_CACHE_DIR, OUTPUT_AD_CACHE_DIR]:
        path = os.path.join(d, fname)
        df.to_csv(path, index=False, encoding="utf-8-sig")
    logger(f"💾 AD cache saved: {fname} ({len(df)} records)")
    return fname

def _find_latest_ad_cache():
    for d in [INPUT_AD_CACHE_DIR, OUTPUT_AD_CACHE_DIR]:
        files = glob.glob(os.path.join(d, "ad_users_*.csv"))
        if files:
            return max(files, key=os.path.getmtime)
    return None

def load_ad_cache():
    path = _find_latest_ad_cache()
    if not path:
        return None, None, "No AD cache found. Run Stage 1 first."
    try:
        df = pd.read_csv(path)
        required = {"email", "displayName"}
        missing = [c for c in required if c not in df.columns]
        if missing:
            return None, path, f"AD cache missing columns: {missing}"
        return df, path, ""
    except Exception as e:
        return None, path, f"Failed to load AD cache: {e}"

def build_identity_index(df):
    email_set = set()
    name_to_emails = {}
    if df is None:
        return email_set, name_to_emails
    for _, row in df.iterrows():
        email = str(row.get("email", "")).strip().lower()
        if not email or email == "nan":
            continue
        email_set.add(email)
        name = normalize_person_name(row.get("displayName", ""))
        if name:
            name_to_emails.setdefault(name, set()).add(email)
    return email_set, name_to_emails


# ============================================
# IDENTITY RESOLUTION (SHARED)
# ============================================

def normalize_person_name(name):
    if pd.isna(name):
        return ""
    return re.sub(r'\s+', ' ', str(name).strip().lower())

def is_email_valid(email):
    return bool(EMAIL_PATTERN.match(str(email).strip().lower()))

def correct_email_domain(email):
    email = str(email).strip().lower()
    for wrong, correct in KNOWN_DOMAIN_CORRECTIONS.items():
        if email.endswith(f"@{wrong}"):
            return email.replace(f"@{wrong}", f"@{correct}"), True
    return email, False

def resolve_identity(value, ad_email_set, ad_name_map):
    """
    Resolve a reviewer/person value to an AD email.
    Returns (ok: bool, canonical_email: str, error: str)
    """
    if pd.isna(value):
        return False, "", "Value is blank/NaN"
    raw = str(value).strip()
    if not raw:
        return False, "", "Value is blank"
    if any(ch in raw for ch in BRACKET_CHARS):
        return False, "", "Value contains bracket characters"
    lower_val = raw.lower()
    if "@" in lower_val:
        if not is_email_valid(lower_val):
            return False, "", "Email format is invalid"
        if lower_val not in ad_email_set:
            corrected, was_corrected = correct_email_domain(lower_val)
            if was_corrected and corrected in ad_email_set:
                return True, corrected, ""
            return False, "", "Email not found in AD"
        return True, lower_val, ""
    normalized = normalize_person_name(raw)
    matches = sorted(list(ad_name_map.get(normalized, set())))
    if len(matches) == 0:
        return False, "", "Name not found in AD"
    if len(matches) > 1:
        return False, "", f"Name maps to {len(matches)} AD accounts"
    return True, matches[0], ""

def fuzzy_match_name(name, ad_name_map, threshold=None):
    if threshold is None:
        threshold = AER_FUZZY_THRESHOLD
    if not FUZZY_AVAILABLE or not name:
        return []
    normalized = normalize_person_name(name)
    results = []
    for ad_name, emails in ad_name_map.items():
        score = rf_fuzz.ratio(normalized, ad_name)
        if score >= threshold:
            results.append((ad_name, score, sorted(list(emails))))
    results.sort(key=lambda x: x[1], reverse=True)
    return results


# ============================================
# SHAREPOINT HELPERS (SHARED)
# ============================================

def normalize_sp_host(value):
    raw = str(value or "").strip().lower()
    if raw.startswith("https://"):
        raw = raw[len("https://"):]
    elif raw.startswith("http://"):
        raw = raw[len("http://"):]
    return raw.split("/", 1)[0].strip()

def sp_ensure_folder(drive_id, parent_item_id, folder_name, headers):
    endpoint = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{parent_item_id}/children"
    # Check existing
    list_url = f"{endpoint}?$filter=name eq '{quote(folder_name)}'"
    resp = graph_get(list_url, headers)
    if resp.status_code < 400:
        items = resp.json().get("value", [])
        for item in items:
            if item.get("name") == folder_name and item.get("folder") is not None:
                return item
    # Create
    payload = {"name": folder_name, "folder": {}, "@microsoft.graph.conflictBehavior": "fail"}
    resp = graph_post(endpoint, headers, json_payload=payload)
    if resp.status_code in (200, 201):
        return resp.json()
    if resp.status_code == 409:
        resp2 = graph_get(f"{endpoint}?$top=999", headers)
        if resp2.status_code < 400:
            for item in resp2.json().get("value", []):
                if item.get("name") == folder_name and item.get("folder") is not None:
                    return item
    raise RuntimeError(f"Ensure folder failed ({resp.status_code}): {resp.text[:300]}")

def sp_upload_file(drive_id, parent_item_id, file_path, headers):
    file_name = os.path.basename(file_path)
    file_size = os.path.getsize(file_path)
    if file_size <= 4 * 1024 * 1024:
        url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{parent_item_id}:/{quote(file_name)}:/content"
        with open(file_path, "rb") as f:
            resp = graph_put(url, headers, data=f.read())
        if resp.status_code >= 400:
            raise RuntimeError(f"Upload failed for {file_name} ({resp.status_code})")
    else:
        _sp_upload_large(drive_id, parent_item_id, file_path, headers)

def _sp_upload_large(drive_id, parent_item_id, file_path, headers):
    file_name = os.path.basename(file_path)
    url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{parent_item_id}:/{quote(file_name)}:/createUploadSession"
    payload = {"item": {"@microsoft.graph.conflictBehavior": "replace", "name": file_name}}
    resp = graph_post(url, headers, json_payload=payload)
    if resp.status_code >= 400:
        raise RuntimeError(f"Upload session failed for {file_name}: {resp.text[:300]}")
    upload_url = resp.json().get("uploadUrl")
    if not upload_url:
        raise RuntimeError(f"Upload session URL missing for {file_name}")
    chunk_size = 5 * 1024 * 1024
    file_size = os.path.getsize(file_path)
    with open(file_path, "rb") as f:
        start = 0
        while start < file_size:
            chunk = f.read(chunk_size)
            end = start + len(chunk) - 1
            put_headers = {
                "Content-Length": str(len(chunk)),
                "Content-Range": f"bytes {start}-{end}/{file_size}",
            }
            chunk_resp = requests.put(upload_url, headers=put_headers, data=chunk)
            if chunk_resp.status_code not in (200, 201, 202):
                raise RuntimeError(f"Chunk upload failed for {file_name}: {chunk_resp.text[:300]}")
            start = end + 1

def sp_list_children(drive_id, parent_item_id, headers):
    items = []
    url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{parent_item_id}/children?$top=999"
    while url:
        resp = graph_get(url, headers)
        if resp.status_code >= 400:
            raise RuntimeError(f"List children failed ({resp.status_code}): {resp.text[:300]}")
        data = resp.json()
        items.extend(data.get("value", []))
        url = data.get("@odata.nextLink")
    return items


# ============================================
# SAFE I/O
# ============================================

def atomic_json_save(file_path, data):
    tmp_path = file_path + ".tmp"
    with open(tmp_path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    os.replace(tmp_path, file_path)

def load_json_safe(file_path):
    if not os.path.exists(file_path):
        return {}
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}

def safe_excel_path(base_path):
    if not os.path.exists(base_path):
        return base_path
    try:
        with open(base_path, "a"):
            pass
        return base_path
    except OSError:
        pass
    stem, ext = os.path.splitext(base_path)
    for i in range(1, 100):
        candidate = f"{stem}_{i}{ext}"
        if not os.path.exists(candidate):
            return candidate
    return f"{stem}_{int(time.time())}{ext}"

def sanitize_folder_name(name):
    safe = re.sub(r'[<>:"/\\|?*]', "_", str(name).strip())
    safe = re.sub(r'_+', '_', safe).strip("_ ")
    return safe or "unnamed"

def detect_encoding(file_path):
    if CHARDET_AVAILABLE:
        with open(file_path, "rb") as f:
            raw = f.read(min(os.path.getsize(file_path), 100000))
        result = chardet.detect(raw)
        return result.get("encoding", "utf-8")
    return "utf-8"


# ============================================
# CHECKPOINT MANAGER
# ============================================

class CheckpointManager:
    def __init__(self, checkpoint_dir=None):
        self.dir = checkpoint_dir or CHECKPOINT_DIR
        os.makedirs(self.dir, exist_ok=True)

    def _path(self, stage):
        return os.path.join(self.dir, f"checkpoint_{stage}.json")

    def load(self, stage):
        return load_json_safe(self._path(stage))

    def mark_done(self, stage, key, metadata=None):
        data = self.load(stage)
        data[key] = {"done": True, "ts": datetime.now().isoformat(), **(metadata or {})}
        atomic_json_save(self._path(stage), data)

    def is_done(self, stage, key):
        data = self.load(stage)
        return data.get(key, {}).get("done", False)

    def clear(self, stage):
        path = self._path(stage)
        if os.path.exists(path):
            os.remove(path)

checkpoint_mgr = None


def _apply_runtime_globals(runtime):
    """Bind runtime values back to the legacy module globals."""
    global _RUNTIME
    global aer_logger
    global TENANT_ID, CLIENT_ID, SHAREPOINT_HOST, SITE_NAME, SENDER_EMAIL
    global AER_REVIEW_YEAR, AER_ROOT_PERSON, AER_ORG_DEPTH, AER_FUZZY_THRESHOLD
    global BASE_PATH, AER_EMAIL_TEMPLATE_FOOTER
    global TODAY_STR, HOUR_STR
    global INPUT_AD_CACHE_DIR, INPUT_MAPPING_DIR
    global OUTPUT_BASE_DIR, OUTPUT_LOG_DIR, OUTPUT_AD_CACHE_DIR, OUTPUT_ORGCHART_DIR
    global STAGE2_DIR, STAGE3_DIR, STAGE4_DIR, REPORT_DIR, CACHE_DIR, CHECKPOINT_DIR
    global LOG_FILE, token_mgr, checkpoint_mgr

    config = runtime.config
    paths = runtime.paths

    TENANT_ID = config.tenant_id
    CLIENT_ID = config.client_id
    SHAREPOINT_HOST = config.sharepoint_host
    SITE_NAME = config.site_name
    SENDER_EMAIL = config.sender_email
    AER_REVIEW_YEAR = config.review_year
    AER_ROOT_PERSON = config.root_person
    AER_ORG_DEPTH = config.org_depth
    AER_FUZZY_THRESHOLD = config.fuzzy_threshold
    BASE_PATH = config.base_path
    AER_EMAIL_TEMPLATE_FOOTER = config.email_template_footer

    TODAY_STR = paths.today_str
    HOUR_STR = paths.hour_str
    INPUT_AD_CACHE_DIR = paths.input_ad_cache_dir
    INPUT_MAPPING_DIR = paths.input_mapping_dir
    OUTPUT_BASE_DIR = paths.output_base_dir
    OUTPUT_LOG_DIR = paths.output_log_dir
    OUTPUT_AD_CACHE_DIR = paths.output_ad_cache_dir
    OUTPUT_ORGCHART_DIR = paths.output_orgchart_dir
    STAGE2_DIR = paths.stage2_dir
    STAGE3_DIR = paths.stage3_dir
    STAGE4_DIR = paths.stage4_dir
    REPORT_DIR = paths.report_dir
    CACHE_DIR = paths.cache_dir
    CHECKPOINT_DIR = paths.checkpoint_dir
    LOG_FILE = paths.log_file

    aer_logger = runtime.logger
    token_mgr = runtime.token_manager_cls(TENANT_ID, CLIENT_ID)
    checkpoint_mgr = CheckpointManager(CHECKPOINT_DIR)
    _RUNTIME = runtime
    return runtime


def build_runtime():
    """Create and bind the shared runtime for notebook stages."""
    config = load_config()
    paths = ensure_runtime_dirs(config)
    logger_obj = setup_logger(paths)
    runtime = AppRuntime(
        config=config,
        paths=paths,
        logger=logger_obj,
        token_manager_cls=TokenManager,
    )
    _apply_runtime_globals(runtime)
    logger(f"Common Library initialized (v5.0) | {TODAY_STR}")
    return runtime


def runtime_status_lines(runtime=None):
    """Return the bootstrap status lines shown in the notebook."""
    active_runtime = runtime or _RUNTIME
    if active_runtime is None:
        raise RuntimeError("Runtime has not been initialized. Call build_runtime() first.")
    return [
        "✅ Common Library loaded",
        f"   Today: {active_runtime.paths.today_str} | Review Year: {active_runtime.config.review_year}",
        f"   Log: {active_runtime.paths.log_file}",
        (
            f"   Fuzzy: {'✅' if FUZZY_AVAILABLE else '❌'} | "
            f"openpyxl: {'✅' if OPENPYXL_AVAILABLE else '❌'} | "
            f"Win32COM: {'✅' if WIN32COM_AVAILABLE else '❌'}"
        ),
        f"   Chardet: {'✅' if CHARDET_AVAILABLE else '❌'} | Tkinter: {'✅' if TK_AVAILABLE else '❌'}",
    ]


def _inject_notebook_globals(namespace, runtime=None):
    """Populate notebook globals with the shared module symbols."""
    active_runtime = runtime or _RUNTIME
    if active_runtime is None:
        raise RuntimeError("Runtime has not been initialized. Call build_runtime() first.")
    if _RUNTIME is not active_runtime:
        _apply_runtime_globals(active_runtime)

    namespace["app_runtime"] = active_runtime
    for name, value in list(globals().items()):
        if name.startswith("_"):
            continue
        namespace[name] = value


# ============================================
# CROSS-PERIOD DIFF HELPER
# ============================================

def find_previous_ad_cache():
    """Find the second-latest AD cache (previous period)."""
    all_files = []
    for d in [INPUT_AD_CACHE_DIR, OUTPUT_AD_CACHE_DIR]:
        all_files.extend(glob.glob(os.path.join(d, "ad_users_*.csv")))
    if len(all_files) < 2:
        return None
    sorted_files = sorted(set(all_files), key=os.path.getmtime, reverse=True)
    return sorted_files[1] if len(sorted_files) >= 2 else None

def compute_diff(current_df, previous_path):
    """Compare current AD snapshot with previous, return diff column."""
    if not previous_path or not os.path.exists(previous_path):
        return pd.Series("FIRST_RUN", index=current_df.index)
    try:
        prev_df = pd.read_csv(previous_path)
    except Exception:
        return pd.Series("FIRST_RUN", index=current_df.index)

    prev_emails = set(prev_df["email"].dropna().str.strip().str.lower())
    prev_lookup = {}
    for _, row in prev_df.iterrows():
        email = str(row.get("email", "")).strip().lower()
        if email and email != "nan":
            prev_lookup[email] = {
                "department": str(row.get("department", "")).strip().lower(),
                "jobTitle": str(row.get("jobTitle", "")).strip().lower(),
            }

    changes = []
    for _, row in current_df.iterrows():
        email = str(row.get("email", "")).strip().lower()
        if not email or email == "nan":
            changes.append("UNKNOWN")
            continue
        if email not in prev_emails:
            changes.append("NEW")
            continue
        prev_info = prev_lookup.get(email, {})
        curr_dept = str(row.get("department", "")).strip().lower()
        curr_title = str(row.get("jobTitle", "")).strip().lower()
        if prev_info.get("department", "") != curr_dept:
            changes.append("DEPT_CHANGED")
        elif prev_info.get("jobTitle", "") != curr_title:
            changes.append("TITLE_CHANGED")
        else:
            changes.append("UNCHANGED")
    return pd.Series(changes, index=current_df.index)


# ============================================
# EXCEL FORMATTING HELPER
# ============================================

def format_export_excel(file_path, audit_col_name="Audit Log"):
    if not OPENPYXL_AVAILABLE:
        return
    wb = load_workbook(file_path)
    ws = wb.active
    header_to_col = {}
    for col_idx in range(1, ws.max_column + 1):
        hdr = ws.cell(row=1, column=col_idx).value
        hdr_txt = str(hdr).strip() if hdr is not None else ""
        if hdr_txt:
            header_to_col[hdr_txt] = col_idx
        max_len = len(hdr_txt)
        for row_idx in range(2, min(ws.max_row + 1, 200)):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val is None:
                continue
            lines = str(cell_val).splitlines() or [str(cell_val)]
            max_len = max(max_len, max(len(line) for line in lines))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(10, max_len + 2), 80)
    audit_col = header_to_col.get(audit_col_name)
    if audit_col:
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=audit_col)
            txt = "" if cell.value is None else str(cell.value)
            line_count = max(1, txt.count("\n") + 1)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            current_height = ws.row_dimensions[row_idx].height or 15
            ws.row_dimensions[row_idx].height = max(current_height, min(15 * line_count, 300))
    wb.save(file_path)
