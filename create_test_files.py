"""
建立測試用的 Excel 檔案（純 openpyxl，不需要 pandas/numpy）
執行後將產生的資料夾上傳到 SharePoint
"""
from openpyxl import Workbook
import os

# 測試資料: (ID, Employee, Access Type, System, Reviewer, Response, Details)
test_data = [
    (1001, "Employee_1", "Read", "SAP", "Joe Petti", "Approved", "No changes needed"),
    (1002, "Employee_2", "Write", "Oracle", "Joe Petti", "Approved", "Maintain current access"),
    (1003, "Employee_3", "Admin", "AWS", "Joe Petti", None, None),  # 缺漏
    (1004, "Employee_4", "Read", "SAP", "Joe Petti", "", ""),  # 缺漏（空字串）
    (1005, "Employee_5", "Write", "Oracle", "Jane Smith", "Rejected", "Remove access - employee left"),
    (1006, "Employee_6", "Read", "SAP", "Jane Smith", "Approved", "Keep read-only access"),
    (1007, "Employee_7", "Admin", "AWS", "Jane Smith", None, None),  # 缺漏
    (1008, "Employee_8", "Read", "Oracle", "Jane Smith", "Approved", "Verified"),
    (1009, "Employee_9", "Write", "SAP", "Jane Smith", "Approved", "OK"),
    (1010, "Employee_10", "Read", "AWS", "Bob O'Brien", "Approved", "Access verified"),
    (1011, "Employee_11", "Write", "Oracle", "Bob O'Brien", "Approved", "No issues"),
    (1012, "Employee_12", "Admin", "SAP", "Bob O'Brien", "Approved", "Confirmed"),
]

headers = ["ID", "Employee Name", "Access Type", "System", "Reviewer", "Reviewer's Response", "Details of Access change"]

# 取得所有審核人
reviewers = list(set(row[4] for row in test_data))
output_dir = "test_data"

print("測試資料概覽:")
print("-" * 60)
for reviewer in sorted(reviewers):
    rows = [r for r in test_data if r[4] == reviewer]
    missing = sum(1 for r in rows if r[5] is None or r[5] == "")
    print(f"  {reviewer}: {len(rows)} 筆, {missing} 筆缺漏")
print("-" * 60)
print()

# 為每個審核人建立獨立的 Excel 檔案
for reviewer in reviewers:
    folder_path = os.path.join(output_dir, reviewer)
    os.makedirs(folder_path, exist_ok=True)

    file_name = f"{reviewer}_Access_Review_2025Q4.xlsx"
    file_path = os.path.join(folder_path, file_name)

    wb = Workbook()
    ws = wb.active
    ws.title = "Access Review"

    # 寫入標題
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    # 寫入所有資料
    for row_idx, row_data in enumerate(test_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # 設定 AutoFilter
    ws.auto_filter.ref = f"A1:G{len(test_data) + 1}"

    # 隱藏不屬於該審核人的行（模擬 filtered view）
    for row_idx, row_data in enumerate(test_data, 2):
        if row_data[4] != reviewer:
            ws.row_dimensions[row_idx].hidden = True

    wb.save(file_path)

    # 統計
    reviewer_rows = [r for r in test_data if r[4] == reviewer]
    missing = sum(1 for r in reviewer_rows if r[5] is None or r[5] == "")

    print(f"✓ {reviewer}:")
    print(f"  檔案: {file_path}")
    print(f"  記錄: {len(reviewer_rows)} 筆, 缺漏: {missing} 筆")
    print()

print("=" * 60)
print("測試資料建立完成！")
print()
print("請將以下資料夾上傳到 SharePoint:")
for reviewer in sorted(reviewers):
    print(f"  test_data/{reviewer}/")
print()
print("上傳目標路徑:")
print("  /sites/aer/Shared Documents/2025 Entitlement Review/Q4/AVIDXCHANGE-TEST/")
